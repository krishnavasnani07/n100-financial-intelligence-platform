"""
Unit and Integration Tests for Peer Comparison Excel & PDF Report Generation.
"""

from __future__ import annotations
import openpyxl
import pytest
from pathlib import Path
from src.config.settings import DB_PATH, OUTPUT_DIR
from src.peer_analysis.comparison import (
    determine_winner,
    generate_executive_summary,
    get_raw_company_ratios,
    generate_peer_comparison_report,
)


def test_determine_winner():
    """Test the KPI winner determination logic."""
    # Test higher-is-better metrics (e.g. ROE)
    assert determine_winner("ROE", 18.5, 12.0, "INFY", "TCS") == "INFY"
    assert determine_winner("ROE", 12.0, 18.5, "INFY", "TCS") == "TCS"
    assert determine_winner("ROE", 15.0, 15.0, "INFY", "TCS") == "Tie"

    # Test lower-is-better metrics (Debt to Equity)
    assert determine_winner("Debt to Equity", 0.1, 0.5, "INFY", "TCS") == "INFY"
    assert determine_winner("Debt to Equity", 0.8, 0.2, "INFY", "TCS") == "TCS"
    assert determine_winner("Debt to Equity", 0.0, 0.0, "INFY", "TCS") == "Tie"

    # Test missing value handling
    assert determine_winner("ROE", None, 15.0, "INFY", "TCS") == "TCS"
    assert determine_winner("ROE", 15.0, None, "INFY", "TCS") == "INFY"


def test_generate_executive_summary():
    """Test that the executive summary is correctly and dynamically formatted."""
    a_data = {
        "ROE": 25.0,
        "ROCE": 30.0,
        "Revenue CAGR": 12.0,
        "PAT CAGR": 15.0,
        "Operating Margin": 24.0,
        "Debt to Equity": 0.0,
        "Composite Quality Score": 85.0,
    }
    b_data = {
        "ROE": 18.0,
        "ROCE": 20.0,
        "Revenue CAGR": 15.0,
        "PAT CAGR": 18.0,
        "Operating Margin": 20.0,
        "Debt to Equity": 0.1,
        "Composite Quality Score": 72.0,
    }

    summary = generate_executive_summary("INFY", "TCS", a_data, b_data)

    assert "INFY outperforms TCS" in summary
    assert "Composite Quality Score" in summary
    assert "growth metrics" in summary
    assert "profitability" in summary


def test_get_raw_company_ratios():
    """Test retrieving raw company metrics from the database."""
    ratios = get_raw_company_ratios("TCS")
    assert ratios is not None
    assert ratios["Company"] == "TCS"
    assert "ROE" in ratios
    assert "Debt to Equity" in ratios


def test_get_raw_company_ratios_invalid():
    """Test that querying a non-existent company raises ValueError."""
    with pytest.raises(ValueError, match="not found in database"):
        get_raw_company_ratios("XYZ_NON_EXISTENT")


def test_generate_peer_comparison_report(tmp_path):
    """Test end-to-end Excel and PDF report generation."""
    # Run the report generator
    excel_path, pdf_path = generate_peer_comparison_report("INFY", "TCS")

    # Assert files exist and are populated
    assert excel_path.exists()
    assert pdf_path.exists()
    assert excel_path.stat().st_size > 0
    assert pdf_path.stat().st_size > 0

    # Read the excel workbook to verify sheet properties
    wb = openpyxl.load_workbook(excel_path)
    assert "Peer Comparison" in wb.sheetnames
    ws = wb["Peer Comparison"]

    # Assert headers and key cells are correct
    assert ws.cell(row=1, column=1).value == "PEER COMPARISON REPORT"
    assert "INFY" in ws.cell(row=2, column=1).value
    assert "TCS" in ws.cell(row=2, column=1).value

    # Verify that the radar chart image is attached to the worksheet
    assert len(ws._images) > 0
