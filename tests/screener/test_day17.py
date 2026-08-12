import pytest
import pandas as pd
import sqlite3
from pathlib import Path
import openpyxl

from src.config.settings import DB_PATH, OUTPUT_DIR
from src.screener.ranking import winsorize_and_scale, calculate_rankings
from src.screener.exporter import generate_reports


def test_winsorize_and_scale_higher_is_better():
    """Test winsorisation and normalization scaling where higher values are better."""
    # Using 11 points: 0.0, 1.0, 2.0, 3.0, 4.0, 5.0, 6.0, 7.0, 8.0, 9.0, 10.0
    # p10 is exactly 1.0, p90 is exactly 9.0
    series = pd.Series([0.0, 1.0, 2.0, 3.0, 4.0, 5.0, 6.0, 7.0, 8.0, 9.0, 10.0])
    scores = winsorize_and_scale(series, lower_is_better=False)

    assert (
        scores.iloc[0] == 0.0
    )  # Outlier 0.0 winsorized to p10 (1.0), which scales to 0.0
    assert scores.iloc[1] == 0.0  # 1.0 is p10, scales to 0.0
    assert (
        scores.iloc[10] == 100.0
    )  # Outlier 10.0 winsorized to p90 (9.0), which scales to 100.0
    assert scores.iloc[9] == 100.0  # 9.0 is p90, scales to 100.0
    # Median is 5.0, scales to 100 * (5.0 - 1.0)/(9.0 - 1.0) = 50.0
    assert scores.iloc[5] == 50.0


def test_winsorize_and_scale_lower_is_better():
    """Test winsorisation and normalization scaling where lower values are better (e.g. D/E)."""
    series = pd.Series([0.0, 1.0, 2.0, 3.0, 4.0, 5.0, 6.0, 7.0, 8.0, 9.0, 10.0])
    scores = winsorize_and_scale(series, lower_is_better=True)

    assert scores.iloc[0] == 100.0  # 0.0 winsorized to p10 (1.0), scales to 100.0
    assert scores.iloc[1] == 100.0  # 1.0 is p10, scales to 100.0
    assert scores.iloc[10] == 0.0  # 10.0 winsorized to p90 (9.0), scales to 0.0
    assert scores.iloc[9] == 0.0  # 9.0 is p90, scales to 0.0
    # Median is 5.0, scales to 100 * (9.0 - 5.0)/(9.0 - 1.0) = 50.0
    assert scores.iloc[5] == 50.0


def test_winsorize_and_scale_all_equal():
    """Test winsorisation and scaling when all values in the series are equal."""
    # Scenario A: positive value
    series_pos = pd.Series([5.0, 5.0, 5.0])
    scores_pos = winsorize_and_scale(series_pos)
    assert (scores_pos == 100.0).all()

    # Scenario B: zero/negative value
    series_neg = pd.Series([0.0, 0.0, 0.0])
    scores_neg = winsorize_and_scale(series_neg)
    assert (scores_neg == 0.0).all()


def test_calculate_rankings_execution():
    """Test calculation of composite scores, rankings, and SQLite database update."""
    df_ranked = calculate_rankings(DB_PATH)

    assert isinstance(df_ranked, pd.DataFrame)
    assert len(df_ranked) > 0

    # Check expected columns from ranking engine
    required_cols = [
        "company_id",
        "overall_rank",
        "composite_quality_score",
        "fcf_cagr_5yr",
        "cfo_pat_ratio",
        "fcf_positive",
    ]
    for col in required_cols:
        assert col in df_ranked.columns

    # Check score ranges (0 to 100)
    assert df_ranked["composite_quality_score"].min() >= 0.0
    assert df_ranked["composite_quality_score"].max() <= 100.0

    # Check rank ordering
    # Ranks should be dense, sorted ascending (Rank 1, 2, 3...)
    assert df_ranked["overall_rank"].iloc[0] == 1
    assert (df_ranked["overall_rank"].diff().dropna() >= 0).all()

    # Verify that SQLite database was updated
    conn = sqlite3.connect(DB_PATH)
    try:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT DISTINCT composite_quality_score FROM financial_ratios WHERE year = ?",
            (df_ranked["year"].iloc[0],),
        )
        db_scores = [r[0] for r in cursor.fetchall()]
        assert len(db_scores) > 1  # Verify that scores are populated and distinct
        assert all(s is not None for s in db_scores)
    finally:
        conn.close()


def test_excel_report_generation():
    """Test that generate_reports creates the styled Excel sheets and individual CSVs."""
    # Execute generation
    generate_reports()

    # Verify file existence
    excel_path_reports = OUTPUT_DIR / "reports" / "screener_output.xlsx"
    excel_path_root = OUTPUT_DIR / "screener_output.xlsx"
    assert excel_path_reports.exists()
    assert excel_path_root.exists()

    # Load and inspect excel sheets
    wb = openpyxl.load_workbook(excel_path_reports)
    expected_sheets = [
        "Summary",
        "Rankings",
        "Quality Compounder",
        "Value Pick",
        "Growth Accelerator",
        "Dividend Champion",
        "Debt-Free Blue Chip",
        "Turnaround Watch",
    ]
    for sheet in expected_sheets:
        assert sheet in wb.sheetnames, f"Missing sheet: {sheet}"

    # Inspect rankings sheet columns
    ws_rankings = wb["Rankings"]
    assert ws_rankings.max_row > 1
    assert ws_rankings.max_column >= 20

    # Verify CSV exports
    csv_dir = OUTPUT_DIR / "csv"
    assert csv_dir.exists()

    expected_csvs = [
        "rankings.csv",
        "quality_compounder.csv",
        "value_pick.csv",
        "growth_accelerator.csv",
        "dividend_champion.csv",
        "debt_free_blue_chip.csv",
        "turnaround_watch.csv",
    ]
    for csv_file in expected_csvs:
        assert (csv_dir / csv_file).exists(), f"Missing CSV: {csv_file}"
