"""
Excel Exporter for Screener Outputs using openpyxl.
Generates output/reports/screener_output.xlsx with professional styling and conditional formatting.
"""

from __future__ import annotations

import datetime
import shutil
import sqlite3
from pathlib import Path
from typing import Any, Dict, List

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.config.settings import DB_PATH, OUTPUT_DIR
from src.screener.presets import run_preset
from src.screener.ranking import calculate_rankings

# Columns to include in Rankings and Preset sheets
COLUMNS_DEFINITION = [
    {"field": "overall_rank", "header": "Rank", "format": "0"},
    {"field": "company_id", "header": "Company ID", "format": "@"},
    {"field": "company_name", "header": "Company Name", "format": "@"},
    {"field": "sector", "header": "Sector", "format": "@"},
    {
        "field": "composite_quality_score",
        "header": "Composite Quality Score",
        "format": "0.00",
    },
    {"field": "return_on_equity_pct", "header": "ROE (%)", "format": "0.0"},
    {"field": "return_on_capital_employed_pct", "header": "ROCE (%)", "format": "0.0"},
    {
        "field": "net_profit_margin_pct",
        "header": "Net Profit Margin (%)",
        "format": "0.0",
    },
    {
        "field": "operating_profit_margin_pct",
        "header": "Operating Margin (%)",
        "format": "0.0",
    },
    {"field": "revenue_cagr_5yr", "header": "Revenue CAGR 5Y (%)", "format": "0.0"},
    {"field": "pat_cagr_5yr", "header": "PAT CAGR 5Y (%)", "format": "0.0"},
    {"field": "fcf_cagr_5yr", "header": "FCF CAGR 5Y (%)", "format": "0.0"},
    {"field": "cfo_pat_ratio", "header": "CFO/PAT Ratio", "format": "0.00"},
    {"field": "free_cash_flow_cr", "header": "FCF (Cr)", "format": "#,##0.0"},
    {"field": "cash_from_operations_cr", "header": "CFO (Cr)", "format": "#,##0.0"},
    {"field": "debt_to_equity", "header": "Debt-to-Equity", "format": "0.00"},
    {"field": "interest_coverage", "header": "Interest Coverage", "format": "0.00"},
    {"field": "asset_turnover", "header": "Asset Turnover", "format": "0.00"},
    {"field": "earnings_per_share", "header": "EPS", "format": "0.00"},
    {"field": "book_value_per_share", "header": "BV per Share", "format": "0.00"},
    {"field": "dividend_yield", "header": "Dividend Yield (%)", "format": "0.0"},
    {
        "field": "dividend_payout_ratio_pct",
        "header": "Dividend Payout (%)",
        "format": "0.0",
    },
    {"field": "pe", "header": "P/E", "format": "0.00"},
    {"field": "pb", "header": "P/B", "format": "0.00"},
    {"field": "sales", "header": "Revenue (Cr)", "format": "#,##0.0"},
    {"field": "revenue_cagr_3yr", "header": "Revenue CAGR 3Y (%)", "format": "0.0"},
    {"field": "de_declining_yoy", "header": "YoY Debt Declining", "format": "@"},
]


def apply_threshold_cell_fill(cell: Any, is_pass: bool) -> None:
    """Applies a professional green/red fill with matching font colors to cells."""
    if is_pass:
        cell.fill = PatternFill(
            start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"
        )
        cell.font = Font(name="Calibri", size=11, color="375623")
    else:
        cell.fill = PatternFill(
            start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"
        )
        cell.font = Font(name="Calibri", size=11, color="C00000")


def apply_conditional_coloring(ws: Any, df: pd.DataFrame, preset_name: str) -> None:
    """Applies color coding to cells that have thresholds in the given preset."""
    col_indices = {col["field"]: idx + 1 for idx, col in enumerate(COLUMNS_DEFINITION)}

    for idx, row in df.iterrows():
        r_idx = idx + 2  # data starts at row 2
        sector = str(row.get("sector", "")).strip().lower()

        # Helper to safely extract float
        def _get_val(field: str) -> float:
            val = row.get(field)
            if val is None or pd.isnull(val):
                return 0.0
            try:
                return float(val)
            except (ValueError, TypeError):
                return 0.0

        if preset_name == "Quality Compounder":
            # ROE > 15%, D/E < 1.0, FCF > 0, Revenue CAGR 5yr > 10%
            apply_threshold_cell_fill(
                ws.cell(row=r_idx, column=col_indices["return_on_equity_pct"]),
                _get_val("return_on_equity_pct") > 15.0,
            )
            apply_threshold_cell_fill(
                ws.cell(row=r_idx, column=col_indices["debt_to_equity"]),
                sector == "financials" or _get_val("debt_to_equity") < 1.0,
            )
            apply_threshold_cell_fill(
                ws.cell(row=r_idx, column=col_indices["free_cash_flow_cr"]),
                _get_val("free_cash_flow_cr") > 0.0,
            )
            apply_threshold_cell_fill(
                ws.cell(row=r_idx, column=col_indices["revenue_cagr_5yr"]),
                _get_val("revenue_cagr_5yr") > 10.0,
            )

        elif preset_name == "Value Pick":
            # P/E < 20, P/B < 3.0, D/E < 2.0, Dividend Yield > 1%
            apply_threshold_cell_fill(
                ws.cell(row=r_idx, column=col_indices["pe"]), _get_val("pe") < 20.0
            )
            apply_threshold_cell_fill(
                ws.cell(row=r_idx, column=col_indices["pb"]), _get_val("pb") < 3.0
            )
            apply_threshold_cell_fill(
                ws.cell(row=r_idx, column=col_indices["debt_to_equity"]),
                sector == "financials" or _get_val("debt_to_equity") < 2.0,
            )
            apply_threshold_cell_fill(
                ws.cell(row=r_idx, column=col_indices["dividend_yield"]),
                _get_val("dividend_yield") > 1.0,
            )

        elif preset_name == "Growth Accelerator":
            # PAT CAGR 5yr > 20%, Revenue CAGR 5yr > 15%, D/E < 2.0
            apply_threshold_cell_fill(
                ws.cell(row=r_idx, column=col_indices["pat_cagr_5yr"]),
                _get_val("pat_cagr_5yr") > 20.0,
            )
            apply_threshold_cell_fill(
                ws.cell(row=r_idx, column=col_indices["revenue_cagr_5yr"]),
                _get_val("revenue_cagr_5yr") > 15.0,
            )
            apply_threshold_cell_fill(
                ws.cell(row=r_idx, column=col_indices["debt_to_equity"]),
                sector == "financials" or _get_val("debt_to_equity") < 2.0,
            )

        elif preset_name == "Dividend Champion":
            # Dividend Yield > 2%, Dividend Payout < 80%, FCF > 0
            apply_threshold_cell_fill(
                ws.cell(row=r_idx, column=col_indices["dividend_yield"]),
                _get_val("dividend_yield") > 2.0,
            )
            apply_threshold_cell_fill(
                ws.cell(row=r_idx, column=col_indices["dividend_payout_ratio_pct"]),
                _get_val("dividend_payout_ratio_pct") < 80.0,
            )
            apply_threshold_cell_fill(
                ws.cell(row=r_idx, column=col_indices["free_cash_flow_cr"]),
                _get_val("free_cash_flow_cr") > 0.0,
            )

        elif preset_name == "Debt-Free Blue Chip":
            # D/E = 0, ROE > 12%, Revenue > 5000 Crore
            apply_threshold_cell_fill(
                ws.cell(row=r_idx, column=col_indices["debt_to_equity"]),
                sector == "financials" or _get_val("debt_to_equity") == 0.0,
            )
            apply_threshold_cell_fill(
                ws.cell(row=r_idx, column=col_indices["return_on_equity_pct"]),
                _get_val("return_on_equity_pct") > 12.0,
            )
            apply_threshold_cell_fill(
                ws.cell(row=r_idx, column=col_indices["sales"]),
                _get_val("sales") > 5000.0,
            )

        elif preset_name == "Turnaround Watch":
            # Revenue CAGR 3yr > 10%, FCF positive, D/E declining YoY
            apply_threshold_cell_fill(
                ws.cell(row=r_idx, column=col_indices["revenue_cagr_3yr"]),
                _get_val("revenue_cagr_3yr") > 10.0,
            )
            apply_threshold_cell_fill(
                ws.cell(row=r_idx, column=col_indices["free_cash_flow_cr"]),
                _get_val("free_cash_flow_cr") > 0.0,
            )

            de_dec = row.get("de_declining_yoy")
            is_declining = (
                de_dec == True
                or de_dec == 1
                or str(de_dec).strip().lower() in ["true", "1"]
            )
            apply_threshold_cell_fill(
                ws.cell(row=r_idx, column=col_indices["de_declining_yoy"]), is_declining
            )


def format_data_sheet(ws: Any, col_definitions: List[Dict[str, Any]]) -> None:
    """Applies premium formatting (navy headers, alternating borders, alignment, etc.) to worksheets."""
    header_fill = PatternFill(
        start_color="1B365D", end_color="1B365D", fill_type="solid"
    )
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    # Format header row
    ws.row_dimensions[1].height = 28
    for col_idx, col in enumerate(col_definitions):
        cell = ws.cell(row=1, column=col_idx + 1)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border

    # Format data rows
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 20
        # Alternate zebra striping (white & light blue/gray background)
        row_fill = PatternFill(
            start_color="F9FBFD" if row % 2 == 0 else "FFFFFF",
            end_color="F9FBFD" if row % 2 == 0 else "FFFFFF",
            fill_type="solid",
        )

        for col_idx, col in enumerate(col_definitions):
            cell = ws.cell(row=row, column=col_idx + 1)
            cell.border = thin_border

            # Apply row fill only if it was not colored by a specific pass/fail threshold
            if cell.fill.fill_type is None:
                cell.fill = row_fill
                cell.font = Font(name="Calibri", size=11, color="000000")

            # Alignment rules
            if col["format"] == "@":
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif col["format"] == "0":
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="right", vertical="center")

            # Apply number format
            if col["format"] and col["format"] != "@":
                cell.number_format = col["format"]

    # Auto-adjust column widths based on values
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or "")
            if cell.row == 1:
                # Truncate headers for length calculation to keep columns reasonable
                max_len = max(max_len, min(14, len(val_str)))
            else:
                max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Add auto-filters
    last_col_letter = get_column_letter(len(col_definitions))
    ws.auto_filter.ref = f"A1:{last_col_letter}{ws.max_row}"


def populate_data_sheet(
    ws: Any, df: pd.DataFrame, col_definitions: List[Dict[str, Any]]
) -> None:
    """Populates rows of a sheet with data based on column field names."""
    # Write headers
    for col_idx, col in enumerate(col_definitions):
        ws.cell(row=1, column=col_idx + 1, value=col["header"])

    # Write data rows
    for r_idx, (_, row) in enumerate(df.iterrows()):
        for c_idx, col in enumerate(col_definitions):
            field = col["field"]
            val = row.get(field)
            # Replace nan with None for Excel
            if pd.isnull(val):
                val = None
            ws.cell(row=r_idx + 2, column=c_idx + 1, value=val)


def create_summary_sheet(
    wb: openpyxl.Workbook,
    rankings_df: pd.DataFrame,
    preset_results: Dict[str, pd.DataFrame],
) -> None:
    """Creates a beautiful, dashboard-like Summary sheet as the first tab in the workbook."""
    ws = wb.active
    ws.title = "Summary"
    ws.views.sheetView[0].showGridLines = True

    # Title block
    ws.cell(row=1, column=1, value="NIFTY 100 FINANCIAL INTELLIGENCE PLATFORM").font = (
        Font(name="Calibri", size=16, bold=True, color="1B365D")
    )
    ws.cell(
        row=2, column=1, value="Screener results & Composite Quality Rankings Summary"
    ).font = Font(name="Calibri", size=11, italic=True, color="595959")
    ws.cell(
        row=3,
        column=1,
        value=f"Report Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
    ).font = Font(name="Calibri", size=9, color="7F7F7F")

    # 1. Universe Quality Statistics
    ws.cell(row=5, column=1, value="Universe Quality Summary Statistics").font = Font(
        name="Calibri", size=13, bold=True, color="1B365D"
    )

    stats = [
        ("Total Companies Analyzed", len(rankings_df), "0"),
        (
            "Average Quality Score",
            rankings_df["composite_quality_score"].mean(),
            "0.00",
        ),
        ("Highest Quality Score", rankings_df["composite_quality_score"].max(), "0.00"),
        ("Lowest Quality Score", rankings_df["composite_quality_score"].min(), "0.00"),
    ]

    header_fill = PatternFill(
        start_color="1B365D", end_color="1B365D", fill_type="solid"
    )
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    ws.cell(row=6, column=1, value="Metric").fill = header_fill
    ws.cell(row=6, column=1).font = header_font
    ws.cell(row=6, column=2, value="Value").fill = header_fill
    ws.cell(row=6, column=2).font = header_font

    for idx, (label, val, fmt) in enumerate(stats):
        r = 7 + idx
        ws.cell(row=r, column=1, value=label).border = thin_border
        cell_val = ws.cell(row=r, column=2, value=val)
        cell_val.border = thin_border
        cell_val.number_format = fmt
        cell_val.alignment = Alignment(horizontal="right")

    # 2. Screener Preset Statistics
    ws.cell(row=13, column=1, value="Screener Preset Statistics").font = Font(
        name="Calibri", size=13, bold=True, color="1B365D"
    )

    preset_headers = [
        "Preset Screener",
        "Description",
        "Companies Passing",
        "Avg. Composite Score",
    ]
    for idx, header in enumerate(preset_headers):
        c = idx + 1
        ws.cell(row=14, column=c, value=header).fill = header_fill
        ws.cell(row=14, column=c).font = header_font

    presets_meta = [
        ("Quality Compounder", "ROE > 15%, D/E < 1.0, FCF > 0, Revenue CAGR 5yr > 10%"),
        ("Value Pick", "P/E < 20, P/B < 3.0, D/E < 2.0, Dividend Yield > 1%"),
        ("Growth Accelerator", "PAT CAGR 5yr > 20%, Revenue CAGR 5yr > 15%, D/E < 2.0"),
        ("Dividend Champion", "Dividend Yield > 2%, Dividend Payout < 80%, FCF > 0"),
        ("Debt-Free Blue Chip", "D/E = 0, ROE > 12%, Revenue > 5000 Crore"),
        ("Turnaround Watch", "Revenue CAGR 3yr > 10%, FCF positive, D/E declining YoY"),
    ]

    for idx, (name, desc) in enumerate(presets_meta):
        r = 15 + idx
        df_res = preset_results[name]
        passing = len(df_res)
        avg_score = df_res["composite_quality_score"].mean() if passing > 0 else 0.0

        ws.cell(row=r, column=1, value=name).border = thin_border
        ws.cell(row=r, column=2, value=desc).border = thin_border

        c_pass = ws.cell(row=r, column=3, value=passing)
        c_pass.border = thin_border
        c_pass.alignment = Alignment(horizontal="center")
        c_pass.number_format = "0"

        c_score = ws.cell(row=r, column=4, value=avg_score)
        c_score.border = thin_border
        c_score.alignment = Alignment(horizontal="right")
        c_score.number_format = "0.00"

    # Auto-adjust column widths for Summary
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row > 4:  # ignore main headers for width
                max_len = max(max_len, len(str(cell.value or "")))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)


def generate_reports() -> None:
    """Generates the full rankings and preset screeners, and exports to Excel and CSVs."""
    # 1. Calculate overall rankings
    print("Calculating quality scores and rankings...")
    rankings_df = calculate_rankings()

    # 2. Run all 6 preset screeners
    presets = [
        "Quality Compounder",
        "Value Pick",
        "Growth Accelerator",
        "Dividend Champion",
        "Debt-Free Blue Chip",
        "Turnaround Watch",
    ]

    preset_results = {}
    csv_dir = OUTPUT_DIR / "csv"
    csv_dir.mkdir(parents=True, exist_ok=True)

    for name in presets:
        # Run preset screener
        df_res = run_preset(name, rankings_df)

        # Sort preset results by composite quality score descending and reset rank
        df_sorted = df_res.sort_values(
            by="composite_quality_score", ascending=False
        ).reset_index(drop=True)
        # Update overall rank in this subset to match the original rank
        preset_results[name] = df_sorted

        # Save preset results to output/csv/
        filename = name.lower().replace(" ", "_").replace("-", "_") + ".csv"
        df_sorted.to_csv(csv_dir / filename, index=False)
        print(f"Exported {name} csv to output/csv/{filename} ({len(df_sorted)} rows)")

    # 3. Create Excel workbook using openpyxl
    wb = openpyxl.Workbook()

    # A. Create Summary sheet (modifies default first sheet)
    create_summary_sheet(wb, rankings_df, preset_results)

    # B. Create Rankings sheet
    ws_rankings = wb.create_sheet(title="Rankings")
    ws_rankings.views.sheetView[0].showGridLines = True
    populate_data_sheet(ws_rankings, rankings_df, COLUMNS_DEFINITION)
    format_data_sheet(ws_rankings, COLUMNS_DEFINITION)

    # C. Create Preset sheets
    for name in presets:
        ws_preset = wb.create_sheet(title=name[:30])  # Excel tabs max length is 31
        ws_preset.views.sheetView[0].showGridLines = True
        df_res = preset_results[name]
        populate_data_sheet(ws_preset, df_res, COLUMNS_DEFINITION)
        apply_conditional_coloring(ws_preset, df_res, name)
        format_data_sheet(ws_preset, COLUMNS_DEFINITION)

    # 4. Save workbook in both reports/ and output/ directories for absolute safety
    reports_dir = OUTPUT_DIR / "reports"
    reports_dir.mkdir(parents=True, exist_ok=True)
    excel_path_reports = reports_dir / "screener_output.xlsx"
    excel_path_root = OUTPUT_DIR / "screener_output.xlsx"

    wb.save(excel_path_reports)
    wb.save(excel_path_root)

    print(f"Workbook successfully exported to: {excel_path_reports}")
    print(f"Workbook successfully exported to: {excel_path_root}")


if __name__ == "__main__":
    generate_reports()
