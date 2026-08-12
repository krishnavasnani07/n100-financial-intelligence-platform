"""
Cash Flow Analytics Engine for Nifty 100 Financial Intelligence Platform.
Computes Free Cash Flow (FCF), CFO Quality Score, CapEx Intensity, FCF Conversion,
and Capital Allocation Classification. Implements Day 31 Health signaling, validations,
and exporters.
"""

from __future__ import annotations

import logging
import math
import os
import sqlite3
import sys
import time
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import numpy as np
import pandas as pd

from src.analytics.cagr import calculate_cagr
from src.analytics.ratio_base import RatioCalculator
from src.config.cashflow_config import (
    CAPEX_ASSET_LIGHT_THRESHOLD,
    CAPEX_INTENSIVE_THRESHOLD,
    CFO_QUALITY_HIGH_THRESHOLD,
    CFO_QUALITY_MODERATE_THRESHOLD,
    DEFAULT_CASHFLOW_PRECISION,
    LABEL_CAPEX_ASSET_LIGHT,
    LABEL_CAPEX_INTENSIVE,
    LABEL_CAPEX_MODERATE,
    LABEL_CFO_ACCRUAL_RISK,
    LABEL_CFO_HIGH,
    LABEL_CFO_MODERATE,
    LABEL_SHAREHOLDER_RETURNS,
    PATTERN_MAP,
)
from src.config.settings import BASE_DIR, DB_PATH, OUTPUT_DIR, RAW_DATA_DIR
from src.utils.helpers import extract_year_int
from src.utils.logger import get_logger

logger = get_logger("cashflow_intelligence")


# Reusable Utility Functions
def safe_divide(
    numerator: Any,
    denominator: Any,
    multiplier: float = 1.0,
    precision: int = DEFAULT_CASHFLOW_PRECISION,
) -> Optional[float]:
    """Safe division helper handling zero, None, and non-numeric types."""
    val, _ = RatioCalculator.safe_divide(
        numerator, denominator, multiplier=multiplier, precision=precision
    )
    return val


def get_sign(value: Any) -> str:
    """Returns '+' if value > 0 else '-'."""
    try:
        val = float(value)
        return "+" if val > 0 else "-"
    except (ValueError, TypeError):
        return "-"


def average_last_n_years(values: List[float], n: int = 5) -> Optional[float]:
    """Computes the average of non-None float values over the last n years."""
    clean_vals = [v for v in values if v is not None and not math.isnan(v)]
    if not clean_vals:
        return None
    recent_vals = clean_vals[-n:]
    return round(sum(recent_vals) / len(recent_vals), DEFAULT_CASHFLOW_PRECISION)


def classify_cfo_quality(ratio: Optional[float]) -> Optional[str]:
    """Classifies CFO Quality ratio into High, Moderate, or Accrual Risk."""
    if ratio is None:
        return None
    if ratio > CFO_QUALITY_HIGH_THRESHOLD:
        return LABEL_CFO_HIGH
    elif ratio >= CFO_QUALITY_MODERATE_THRESHOLD:
        return LABEL_CFO_MODERATE
    else:
        return LABEL_CFO_ACCRUAL_RISK


def classify_capex_intensity(percentage: Optional[float]) -> Optional[str]:
    """Classifies CapEx Intensity percentage into Asset Light, Moderate, or Capital Intensive."""
    if percentage is None:
        return None
    if percentage < CAPEX_ASSET_LIGHT_THRESHOLD:
        return LABEL_CAPEX_ASSET_LIGHT
    elif percentage <= CAPEX_INTENSIVE_THRESHOLD:
        return LABEL_CAPEX_MODERATE
    else:
        return LABEL_CAPEX_INTENSIVE


# Standalone KPI Calculations
def calculate_free_cash_flow(
    operating_activity: Any,
    investing_activity: Any,
    precision: int = DEFAULT_CASHFLOW_PRECISION,
    company_id: str = "UNKNOWN",
    year: str = "UNKNOWN",
) -> Optional[float]:
    """
    Computes Free Cash Flow (FCF) = Operating Cash Flow + Investing Cash Flow.
    """
    try:
        cfo = float(operating_activity) if operating_activity is not None else 0.0
        cfi = float(investing_activity) if investing_activity is not None else 0.0
        fcf = round(cfo + cfi, precision)
        return fcf
    except (ValueError, TypeError):
        return None


def calculate_cfo_quality(
    operating_activity: Any,
    net_profit: Any,
    precision: int = DEFAULT_CASHFLOW_PRECISION,
    company_id: str = "UNKNOWN",
    year: str = "UNKNOWN",
) -> Optional[float]:
    """
    Computes single-period CFO Quality ratio = Operating Cash Flow / Net Profit (PAT).
    """
    try:
        pat = float(net_profit) if net_profit is not None else 0.0
        if pat <= 0:
            return None
    except (ValueError, TypeError):
        return None

    return safe_divide(operating_activity, pat, multiplier=1.0, precision=precision)


def calculate_capex_intensity(
    investing_activity: Any,
    sales: Any,
    precision: int = DEFAULT_CASHFLOW_PRECISION,
    company_id: str = "UNKNOWN",
    year: str = "UNKNOWN",
) -> Optional[float]:
    """
    Computes CapEx Intensity (%) = ABS(Investing Cash Flow) / Sales * 100.
    """
    try:
        s = float(sales) if sales is not None else 0.0
        if s <= 0:
            return None
        cfi = float(investing_activity) if investing_activity is not None else 0.0
        return safe_divide(abs(cfi), s, multiplier=100.0, precision=precision)
    except (ValueError, TypeError):
        return None


def calculate_fcf_conversion(
    free_cash_flow: Any,
    operating_profit: Any,
    precision: int = DEFAULT_CASHFLOW_PRECISION,
    company_id: str = "UNKNOWN",
    year: str = "UNKNOWN",
) -> Optional[float]:
    """
    Computes FCF Conversion = Free Cash Flow / Operating Profit * 100.
    """
    try:
        op = float(operating_profit) if operating_profit is not None else 0.0
        if op <= 0:
            return None
        fcf = float(free_cash_flow) if free_cash_flow is not None else 0.0
        return safe_divide(fcf, op, multiplier=100.0, precision=precision)
    except (ValueError, TypeError):
        return None


def classify_capital_allocation(
    cfo: Any,
    cfi: Any,
    cff: Any,
    cfo_pat_ratio: Optional[float] = None,
    company_id: str = "UNKNOWN",
    year: str = "UNKNOWN",
) -> Tuple[str, str, str, str]:
    """
    Classifies company capital allocation into one of 8 patterns based on cash flow signs.
    """
    cfo_sign = get_sign(cfo)
    cfi_sign = get_sign(cfi)
    cff_sign = get_sign(cff)
    key = (cfo_sign, cfi_sign, cff_sign)

    if key == ("+", "-", "-"):
        if cfo_pat_ratio is not None and cfo_pat_ratio > CFO_QUALITY_HIGH_THRESHOLD:
            label = LABEL_SHAREHOLDER_RETURNS
        else:
            label = PATTERN_MAP[key]
    else:
        label = PATTERN_MAP.get(key, "Mixed")

    return cfo_sign, cfi_sign, cff_sign, label


# Step 2: Data Loaders and Validation
def load_cashflow(db_path: Path) -> pd.DataFrame:
    conn = sqlite3.connect(str(db_path))
    try:
        df = pd.read_sql_query(
            "SELECT company_id, year, operating_activity, investing_activity, financing_activity FROM cashflow",
            conn,
        )
        df["company_id"] = df["company_id"].astype(str).str.strip().str.upper()
        return df
    finally:
        conn.close()


def load_profit_loss(db_path: Path) -> pd.DataFrame:
    conn = sqlite3.connect(str(db_path))
    try:
        df = pd.read_sql_query(
            "SELECT company_id, year, sales, expenses, operating_profit, net_profit FROM profitandloss",
            conn,
        )
        df["company_id"] = df["company_id"].astype(str).str.strip().str.upper()
        return df
    finally:
        conn.close()


def load_balance_sheet(db_path: Path) -> pd.DataFrame:
    conn = sqlite3.connect(str(db_path))
    try:
        df = pd.read_sql_query(
            "SELECT company_id, year, borrowings, total_assets FROM balancesheet", conn
        )
        df["company_id"] = df["company_id"].astype(str).str.strip().str.upper()
        return df
    finally:
        conn.close()


def load_ratios(db_path: Path) -> pd.DataFrame:
    conn = sqlite3.connect(str(db_path))
    try:
        df = pd.read_sql_query(
            "SELECT company_id, year, debt_to_equity, return_on_equity_pct, free_cash_flow_cr FROM financial_ratios",
            conn,
        )
        df["company_id"] = df["company_id"].astype(str).str.strip().str.upper()
        return df
    finally:
        conn.close()


def load_companies(db_path: Path) -> pd.DataFrame:
    conn = sqlite3.connect(str(db_path))
    try:
        df = pd.read_sql_query(
            "SELECT id AS company_id, company_name FROM companies", conn
        )
        df["company_id"] = df["company_id"].astype(str).str.strip().str.upper()
        return df
    finally:
        conn.close()


def load_capital_allocation(filepath: Path) -> pd.DataFrame:
    if not filepath.exists():
        logger.warning(f"capital_allocation.csv not found at {filepath}")
        return pd.DataFrame(columns=["company_id", "year", "pattern_label"])
    df = pd.read_csv(filepath)
    df["company_id"] = df["company_id"].astype(str).str.strip().str.upper()
    return df


def validate_inputs(
    df_cf: pd.DataFrame,
    df_pl: pd.DataFrame,
    df_bs: pd.DataFrame,
    df_ratios: pd.DataFrame,
    df_companies: pd.DataFrame,
    df_alloc: pd.DataFrame,
) -> List[str]:
    warnings = []
    if df_companies.empty:
        warnings.append("Companies table is empty.")
    if df_companies["company_id"].duplicated().any():
        warnings.append("Duplicate company IDs found in companies table.")
    if df_cf.empty:
        warnings.append("Cashflow table is empty.")
    if df_pl.empty:
        warnings.append("Profitandloss table is empty.")
    if df_bs.empty:
        warnings.append("Balancesheet table is empty.")

    # Check missing metrics
    if (
        df_cf[["operating_activity", "investing_activity", "financing_activity"]]
        .isnull()
        .any()
        .any()
    ):
        warnings.append("Missing cash flow metrics (CFO/CFI/CFF) in cashflow data.")
    if df_pl[["sales", "net_profit"]].isnull().any().any():
        warnings.append("Missing sales or net_profit in profitandloss data.")

    return warnings


# Engine Class (Retaining previous code compatibility)
class CashFlowEngine:
    @classmethod
    def compute_period_kpis(
        cls,
        company_id: str,
        year: str,
        operating_activity: Any,
        investing_activity: Any,
        financing_activity: Any,
        sales: Any,
        operating_profit: Any,
        net_profit: Any,
        cfo_pat_5yr_avg: Optional[float] = None,
    ) -> Dict[str, Any]:
        """
        Computes all cash flow KPIs for a single period.

        Parameters:
            company_id (str): ID of the company.
            year (str): Financial year.
            operating_activity (Any): Cash flow from operating activities.
            investing_activity (Any): Cash flow from investing activities.
            financing_activity (Any): Cash flow from financing activities.
            sales (Any): Revenue from sales.
            operating_profit (Any): Operating profit.
            net_profit (Any): Net profit (PAT).
            cfo_pat_5yr_avg (Optional[float]): 5-year average CFO/PAT quality ratio.

        Returns:
            Dict[str, Any]: Dictionary of calculated cash flow KPIs and labels.
        """
        fcf = calculate_free_cash_flow(
            operating_activity, investing_activity, company_id=company_id, year=year
        )
        cfo_qual_period = calculate_cfo_quality(
            operating_activity, net_profit, company_id=company_id, year=year
        )

        eff_cfo_pat = (
            cfo_pat_5yr_avg if cfo_pat_5yr_avg is not None else cfo_qual_period
        )
        cfo_qual_label = classify_cfo_quality(eff_cfo_pat)

        capex_pct = calculate_capex_intensity(
            investing_activity, sales, company_id=company_id, year=year
        )
        capex_label = classify_capex_intensity(capex_pct)

        # Note: Retention of previous logic where FCF conversion was ratio, not pct
        fcf_conv = calculate_fcf_conversion(
            fcf, operating_profit, company_id=company_id, year=year
        )
        if fcf_conv is not None:
            fcf_conv = round(fcf_conv / 100.0, 4)

        cfo_sign, cfi_sign, cff_sign, pattern_label = classify_capital_allocation(
            cfo=operating_activity,
            cfi=investing_activity,
            cff=financing_activity,
            cfo_pat_ratio=eff_cfo_pat,
            company_id=company_id,
            year=year,
        )

        return {
            "company_id": company_id,
            "year": str(year),
            "free_cash_flow": fcf,
            "cfo_quality_period": cfo_qual_period,
            "cfo_quality_5yr_avg": cfo_pat_5yr_avg,
            "cfo_quality_label": cfo_qual_label,
            "capex_intensity_pct": capex_pct,
            "capex_intensity_label": capex_label,
            "fcf_conversion": fcf_conv,
            "cfo_sign": cfo_sign,
            "cfi_sign": cfi_sign,
            "cff_sign": cff_sign,
            "pattern_label": pattern_label,
        }

    @classmethod
    def compute_company_cashflow_kpis(
        cls, company_id: str, df_company: pd.DataFrame
    ) -> List[Dict[str, Any]]:
        """
        Computes cash flow KPIs chronologically for all available years of a company.

        Parameters:
            company_id (str): ID of the company.
            df_company (pd.DataFrame): Dataframe containing historical financial data.

        Returns:
            List[Dict[str, Any]]: Chronological list of period KPIs.
        """
        df_sorted = df_company.sort_values(by="year").copy()
        cfo_qualities = []
        for _, row in df_sorted.iterrows():
            q = calculate_cfo_quality(
                row.get("operating_activity"),
                row.get("net_profit"),
                company_id=company_id,
                year=row.get("year"),
            )
            cfo_qualities.append(q)

        results = []
        for i, (_, row) in enumerate(df_sorted.iterrows()):
            hist_q = cfo_qualities[max(0, i - 4) : i + 1]
            avg_5yr = average_last_n_years(hist_q, n=5)

            res = cls.compute_period_kpis(
                company_id=company_id,
                year=row.get("year"),
                operating_activity=row.get("operating_activity"),
                investing_activity=row.get("investing_activity"),
                financing_activity=row.get("financing_activity"),
                sales=row.get("sales"),
                operating_profit=row.get("operating_profit"),
                net_profit=row.get("net_profit"),
                cfo_pat_5yr_avg=avg_5yr,
            )
            results.append(res)
        return results

    @classmethod
    def export_cashflow_reports(
        cls, results: List[Dict[str, Any]], output_dir: Optional[Path] = None
    ) -> Dict[str, Path]:
        """
        Exports cash flow KPI summaries and capital allocation patterns to CSV.

        Parameters:
            results (List[Dict[str, Any]]): List of period calculation results.
            output_dir (Optional[Path]): Directory where CSVs should be saved.

        Returns:
            Dict[str, Path]: Mapping of report names to exported file paths.
        """
        out_dir = output_dir or (BASE_DIR / "output")
        out_dir.mkdir(parents=True, exist_ok=True)
        df_all = pd.DataFrame(results)

        file_cap_alloc = out_dir / "capital_allocation.csv"
        alloc_cols = [
            "company_id",
            "year",
            "cfo_sign",
            "cfi_sign",
            "cff_sign",
            "pattern_label",
        ]
        df_alloc = (
            df_all[alloc_cols] if not df_all.empty else pd.DataFrame(columns=alloc_cols)
        )
        df_alloc.to_csv(file_cap_alloc, index=False)

        file_summary = out_dir / "cashflow_summary.csv"
        df_all.to_csv(file_summary, index=False)

        file_stats = out_dir / "capital_pattern_statistics.csv"
        if not df_all.empty and "pattern_label" in df_all.columns:
            counts = df_all["pattern_label"].value_counts().reset_index()
            counts.columns = ["pattern_label", "count"]
            counts["percentage"] = (counts["count"] / len(df_all) * 100).round(2)
        else:
            counts = pd.DataFrame(columns=["pattern_label", "count", "percentage"])
        counts.to_csv(file_stats, index=False)

        return {
            "capital_allocation": file_cap_alloc,
            "cashflow_summary": file_summary,
            "pattern_statistics": file_stats,
        }


# Master run function for Day 31 Health Engine
def run_cashflow_intelligence_pipeline(
    db_path: Optional[Path] = None, output_dir: Optional[Path] = None
) -> pd.DataFrame:
    """
    Executes the complete cash flow intelligence analytics pipeline.
    Loads data, runs validations, computes rolling KPIs and allocation patterns,
    and exports styled Excel/CSV sheets.

    Parameters:
        db_path (Optional[Path]): Path to the SQLite database.
        output_dir (Optional[Path]): Path to directory for exporting reports.

    Returns:
        pd.DataFrame: Calculated cash flow intelligence data.
    """
    start_time = time.time()
    db_file = db_path or DB_PATH
    out_dir = output_dir or OUTPUT_DIR
    out_dir.mkdir(parents=True, exist_ok=True)

    logger.info("Initializing Cash Flow Intelligence Pipeline...")

    try:
        # Load datasets
        df_cf = load_cashflow(db_file)
        df_pl = load_profit_loss(db_file)
        df_bs = load_balance_sheet(db_file)
        df_ratios = load_ratios(db_file)
        df_companies = load_companies(db_file)

        cap_alloc_csv = out_dir / "capital_allocation.csv"
        df_alloc = load_capital_allocation(cap_alloc_csv)

        # Validation checks
        warnings = validate_inputs(
            df_cf, df_pl, df_bs, df_ratios, df_companies, df_alloc
        )
        for w in warnings:
            logger.warning(w)

        # Get sector mappings
        conn = sqlite3.connect(str(db_file))
        df_sectors = pd.read_sql_query(
            "SELECT company_id, broad_sector AS sector FROM sectors", conn
        )
        conn.close()
        df_sectors["company_id"] = (
            df_sectors["company_id"].astype(str).str.strip().str.upper()
        )
        sectors_map = dict(zip(df_sectors["company_id"], df_sectors["sector"]))

    except Exception as e:
        logger.error(f"Error loading datasets: {e}", exc_info=True)
        raise

    # Compile company histories chronologically
    df_merged = pd.merge(df_cf, df_pl, on=["company_id", "year"], how="left")
    df_merged = pd.merge(df_merged, df_bs, on=["company_id", "year"], how="left")

    df_merged["year_int"] = df_merged["year"].apply(extract_year_int)
    df_merged = df_merged.dropna(subset=["year_int"])
    df_merged["year_int"] = df_merged["year_int"].astype(int)

    company_histories: Dict[str, List[Dict[str, Any]]] = {}
    for cid, group in df_merged.groupby("company_id"):
        group_sorted = group.sort_values(by="year_int")

        # Precompute FCF for each historical record
        records = []
        for _, row in group_sorted.iterrows():
            cfo = row.get("operating_activity")
            cfi = row.get("investing_activity")
            fcf = calculate_free_cash_flow(cfo, cfi)
            rec = row.to_dict()
            rec["free_cash_flow"] = fcf
            records.append(rec)

        company_histories[cid] = records

    # Load Capital Allocation Latest mappings
    if not df_alloc.empty:
        df_alloc["year_int"] = df_alloc["year"].apply(extract_year_int)
        df_alloc_latest = df_alloc.sort_values(by="year_int").drop_duplicates(
            subset=["company_id"], keep="last"
        )
        alloc_map = dict(
            zip(df_alloc_latest["company_id"], df_alloc_latest["pattern_label"])
        )
    else:
        alloc_map = {}

    # Process indicators for each company
    intelligence_records = []
    distress_alerts = []

    high_quality_count = 0
    moderate_count = 0
    accrual_risk_count = 0

    for idx, comp_row in df_companies.iterrows():
        company_id = str(comp_row["company_id"]).strip().upper()
        company_name = comp_row["company_name"]

        history = company_histories.get(company_id, [])
        sector = sectors_map.get(company_id, "Other")

        if not history:
            logger.warning(
                f"No history found for company: {company_id}. Generating default entry."
            )
            intelligence_records.append(
                {
                    "company_id": company_id,
                    "sector": sector,
                    "cfo_quality_score": None,
                    "cfo_quality_label": None,
                    "capex_intensity_pct": None,
                    "capex_label": None,
                    "fcf_cagr_5yr": None,
                    "fcf_conversion_pct": None,
                    "distress_flag": False,
                    "deleveraging_flag": False,
                    "capital_allocation_label": "Mixed",
                }
            )
            continue

        latest_rec = history[-1]

        # 1. CFO Quality Score & Label (5-year rolling average of CFO / PAT)
        q_scores = []
        for r in history[-5:]:
            cfo_period = r.get("operating_activity")
            pat_period = r.get("net_profit")
            if cfo_period is not None and pat_period is not None and pat_period > 0:
                q_scores.append(cfo_period / pat_period)

        cfo_score = round(sum(q_scores) / len(q_scores), 4) if q_scores else None
        cfo_label = classify_cfo_quality(cfo_score)

        if cfo_label == LABEL_CFO_HIGH:
            high_quality_count += 1
        elif cfo_label == LABEL_CFO_MODERATE:
            moderate_count += 1
        elif cfo_label == LABEL_CFO_ACCRUAL_RISK:
            accrual_risk_count += 1

        # 2. CapEx Intensity
        cfi = latest_rec.get("investing_activity")
        sales = latest_rec.get("sales")
        capex_intensity = None
        if cfi is not None and sales is not None and sales > 0:
            capex_intensity = round((abs(cfi) / sales) * 100, 4)
        capex_label = classify_capex_intensity(capex_intensity)

        # 3. FCF CAGR 5Y (End FCF, Start FCF = latest - 5 years)
        fcf_cagr = None
        latest_year_int = latest_rec["year_int"]
        start_year_int = latest_year_int - 5

        start_fcf = None
        for r in history:
            if r["year_int"] == start_year_int:
                start_fcf = r.get("free_cash_flow")
                break

        if start_fcf is not None and latest_rec.get("free_cash_flow") is not None:
            cagr_val, _ = calculate_cagr(
                start_value=start_fcf,
                end_value=latest_rec.get("free_cash_flow"),
                years=5,
                company_id=company_id,
                metric_name="FCF",
            )
            fcf_cagr = cagr_val

        # 4. FCF Conversion % (FCF / Operating Profit * 100)
        fcf_conv = None
        op = latest_rec.get("operating_profit")
        fcf_latest = latest_rec.get("free_cash_flow")
        if fcf_latest is not None and op is not None and op > 0:
            fcf_conv = round((fcf_latest / op) * 100, 4)

        # 5. Distress flag: CFO < 0 and CFF > 0
        cfo = latest_rec.get("operating_activity")
        cff = latest_rec.get("financing_activity")
        distress = bool(cfo is not None and cff is not None and cfo < 0 and cff > 0)

        # 6. Deleveraging flag: CFF < 0 and YoY Borrowings declined
        borrowings_latest = latest_rec.get("borrowings")
        borrowings_prev = history[-2].get("borrowings") if len(history) >= 2 else None

        deleveraging = bool(
            cff is not None
            and cff < 0
            and borrowings_latest is not None
            and borrowings_prev is not None
            and borrowings_latest < borrowings_prev
        )

        # 7. Capital Allocation Label
        cap_alloc_label = alloc_map.get(company_id, "Mixed")

        # Record output record
        intelligence_records.append(
            {
                "company_id": company_id,
                "sector": sector,
                "cfo_quality_score": cfo_score,
                "cfo_quality_label": cfo_label,
                "capex_intensity_pct": capex_intensity,
                "capex_label": capex_label,
                "fcf_cagr_5yr": fcf_cagr,
                "fcf_conversion_pct": fcf_conv,
                "distress_flag": distress,
                "deleveraging_flag": deleveraging,
                "capital_allocation_label": cap_alloc_label,
            }
        )

        # Record distress alerts
        if distress:
            pat_latest = latest_rec.get("net_profit")
            reason = (
                f"Negative Operating Cash Flow ({cfo:g} Cr) and Positive Financing Cash Flow ({cff:g} Cr). "
                f"Business is consuming operational cash while raising funds."
            )
            distress_alerts.append(
                {
                    "company_id": company_id,
                    "company_name": company_name,
                    "CFO": cfo,
                    "CFF": cff,
                    "Net Profit": pat_latest,
                    "Distress Reason": reason,
                }
            )

    df_intel = pd.DataFrame(intelligence_records)

    # Sort: Sector -> Company
    df_intel = df_intel.sort_values(by=["sector", "company_id"]).reset_index(drop=True)

    # Export Distress Alerts CSV
    df_alerts = pd.DataFrame(distress_alerts)
    alerts_csv_path = out_dir / "distress_alerts.csv"
    if df_alerts.empty:
        # Create empty template
        df_alerts = pd.DataFrame(
            columns=[
                "company_id",
                "company_name",
                "CFO",
                "CFF",
                "Net Profit",
                "Distress Reason",
            ]
        )
    df_alerts.to_csv(alerts_csv_path, index=False)

    # Export styled Excel report
    excel_path = out_dir / "cashflow_intelligence.xlsx"
    export_excel_intelligence(df_intel, excel_path)

    runtime = round(time.time() - start_time, 4)

    # Logs
    logger.info(
        f"Cashflow Intelligence Engine finished in {runtime}s. "
        f"Companies Processed: {len(df_intel)}, "
        f"High Quality: {high_quality_count}, "
        f"Moderate: {moderate_count}, "
        f"Accrual Risk: {accrual_risk_count}, "
        f"Distress Alerts: {len(df_alerts)}."
    )

    # Validation checks (Step 14)
    validation_passed = True
    val_msg = []

    if len(df_intel) != 92:
        val_msg.append(
            f"Validation warning: Processed {len(df_intel)} companies instead of 92."
        )
        validation_passed = False

    if df_intel.empty:
        val_msg.append("Validation Fail: Intelligence dataset is empty.")
        validation_passed = False

    if df_intel.duplicated(subset=["company_id"]).any():
        val_msg.append("Validation Fail: Duplicate company ID rows found in output.")
        validation_passed = False

    if validation_passed:
        logger.info("Validation checks passed successfully.")
    else:
        logger.error(f"Validation checks failed: {val_msg}")

    # Print nice display block
    print("=" * 80)
    print("             CASH FLOW INTELLIGENCE & HEALTH ENGINE              ")
    print("=" * 80)
    print(f"[+] Total Companies Processed : {len(df_intel)}")
    print(f"    High CFO Quality Count    : {high_quality_count}")
    print(f"    Moderate Quality Count    : {moderate_count}")
    print(f"    Accrual Risk Count        : {accrual_risk_count}")
    print(f"    Distress Alerts Generated : {len(df_alerts)}")
    print(f"    Engine Runtime            : {runtime}s")
    print(f"    Output Excel File Path    : {excel_path}")
    print(f"    Output Distress CSV Path  : {alerts_csv_path}")
    print(
        f"    Validation Status         : {'PASSED' if validation_passed else 'FAILED'}"
    )
    if not validation_passed:
        for m in val_msg:
            print(f"    [!] {m}", file=sys.stderr)
    print("-" * 80)

    return df_intel


def export_excel_intelligence(df: pd.DataFrame, output_path: Path):
    """
    Exports styled Excel workbook for cash flow intelligence.
    """
    df_excel = df.copy()

    # Excel print column order and capitalization (9 columns as specified in Day 32)
    excel_cols = {
        "company_id": "company_id",
        "sector": "sector",
        "cfo_quality_label": "CFO Quality",
        "capex_label": "CapEx Label",
        "fcf_cagr_5yr": "FCF CAGR",
        "fcf_conversion_pct": "FCF Conversion",
        "distress_flag": "Distress",
        "deleveraging_flag": "Deleveraging",
        "capital_allocation_label": "Capital Allocation",
    }
    df_excel = df_excel[list(excel_cols.keys())].rename(columns=excel_cols)

    # Convert percentages to decimals so Excel formats them correctly
    for c in ["FCF CAGR", "FCF Conversion"]:
        df_excel[c] = df_excel[c] / 100.0

    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        writer = pd.ExcelWriter(output_path, engine="openpyxl")
        df_excel.to_excel(writer, sheet_name="Cash Flow Health", index=False)

        workbook = writer.book
        worksheet = writer.sheets["Cash Flow Health"]

        # Design system styles
        font_family = "Inter"
        header_fill = PatternFill(
            start_color="1E3A8A", end_color="1E3A8A", fill_type="solid"
        )
        header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
        data_font = Font(name=font_family, size=10)

        # Zebra striping
        zebra_fill = PatternFill(
            start_color="F8FAFC", end_color="F8FAFC", fill_type="solid"
        )

        # Indicator Fills
        green_fill = PatternFill(
            start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"
        )
        green_font = Font(name=font_family, size=10, bold=True, color="166534")

        yellow_fill = PatternFill(
            start_color="FEF08A", end_color="FEF08A", fill_type="solid"
        )
        yellow_font = Font(name=font_family, size=10, bold=True, color="854D0E")

        red_fill = PatternFill(
            start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"
        )
        red_font = Font(name=font_family, size=10, bold=True, color="991B1B")

        thin_border = Border(
            left=Side(style="thin", color="E2E8F0"),
            right=Side(style="thin", color="E2E8F0"),
            top=Side(style="thin", color="E2E8F0"),
            bottom=Side(style="thin", color="E2E8F0"),
        )

        # Header formatting
        for col_idx in range(1, len(df_excel.columns) + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            cell.border = thin_border

        worksheet.row_dimensions[1].height = 28

        # Data formatting
        for row_idx in range(2, len(df_excel) + 2):
            worksheet.row_dimensions[row_idx].height = 20
            is_zebra = row_idx % 2 == 0

            cfo_lbl = worksheet.cell(row=row_idx, column=3).value
            capex_lbl = worksheet.cell(row=row_idx, column=4).value
            distress_val = worksheet.cell(row=row_idx, column=7).value

            for col_idx in range(1, len(df_excel.columns) + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.font = data_font
                cell.border = thin_border

                # Default Zebra
                if is_zebra:
                    cell.fill = zebra_fill

                # Alignments
                if col_idx in [
                    1,
                    2,
                    3,
                    4,
                    7,
                    8,
                    9,
                ]:  # IDs, Labels, Flags, Capital Allocation
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:  # Numeric ratios/scores
                    cell.alignment = Alignment(horizontal="right", vertical="center")

                # Number formatting
                val = cell.value
                if val is not None and isinstance(val, (int, float)):
                    if col_idx in [5, 6]:  # Percentages
                        cell.number_format = "0.00%"

                # Conditional Fills
                if col_idx == 3:  # CFO quality
                    if cfo_lbl == "High" or cfo_lbl == "High Quality":
                        cell.fill = green_fill
                        cell.font = green_font
                    elif cfo_lbl == "Moderate":
                        cell.fill = yellow_fill
                        cell.font = yellow_font
                    elif cfo_lbl == "Accrual Risk":
                        cell.fill = red_fill
                        cell.font = red_font

                elif col_idx == 4:  # CapEx label
                    if capex_lbl == "Asset Light":
                        cell.fill = green_fill
                        cell.font = green_font
                    elif capex_lbl == "Moderate":
                        cell.fill = yellow_fill
                        cell.font = yellow_font
                    elif capex_lbl == "Capital Intensive":
                        cell.fill = red_fill
                        cell.font = red_font

                elif col_idx == 7:  # Distress flag
                    if str(distress_val).strip() == "True":
                        cell.fill = red_fill
                        cell.font = red_font
                    else:
                        cell.fill = green_fill
                        cell.font = green_font

        worksheet.freeze_panes = "A2"

        # Column auto width
        for col in worksheet.columns:
            max_len = 0
            for cell in col:
                val = str(cell.value or "")
                if len(val) > max_len:
                    max_len = len(val)
            col_letter = get_column_letter(col[0].column)
            worksheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

        writer.close()
        logger.info(
            f"Successfully exported styled cashflow intelligence workbook to {output_path}"
        )

    except Exception as e:
        logger.error(
            f"Failed to style cashflow intelligence Excel report: {e}", exc_info=True
        )
        df_excel.to_excel(output_path, sheet_name="Cash Flow Health", index=False)
