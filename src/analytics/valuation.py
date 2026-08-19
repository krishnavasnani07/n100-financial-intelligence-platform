"""
Valuation Engine for Nifty 100 Financial Intelligence Platform.
Computes FCF Yield, Sector Median PE, comparisons, valuation flags, and exports reports.
"""

from __future__ import annotations

import logging
import sqlite3
import sys
import time
from pathlib import Path

import pandas as pd

from src.config.settings import BASE_DIR, DB_PATH, OUTPUT_DIR, RAW_DATA_DIR

# Custom logging to logs/valuation.log and stdout
LOG_FILE = BASE_DIR / "logs" / "valuation.log"
LOG_FILE.parent.mkdir(parents=True, exist_ok=True)

logger = logging.getLogger("valuation")
logger.setLevel(logging.INFO)

# Clear existing handlers if any
if logger.hasHandlers():
    logger.handlers.clear()

formatter = logging.Formatter("%(asctime)s - %(name)s - %(levelname)s - %(message)s")

# Console handler
console_handler = logging.StreamHandler(sys.stdout)
console_handler.setLevel(logging.INFO)
console_handler.setFormatter(formatter)
logger.addHandler(console_handler)

# File handler
try:
    file_handler = logging.FileHandler(LOG_FILE, encoding="utf-8")
    file_handler.setLevel(logging.INFO)
    file_handler.setFormatter(formatter)
    logger.addHandler(file_handler)
except Exception as e:
    console_handler.stream.write(
        f"CRITICAL: Failed to initialize file log handler at {LOG_FILE}: {e}\n"
    )


from src.utils.helpers import extract_year_int


def load_market_cap(filepath: Path | None = None) -> pd.DataFrame:
    """
    Loads raw market cap data from Excel, filters to the latest year for each company,
    computes the 5-year median PE across historical records, and performs validations.
    """
    path = filepath or (RAW_DATA_DIR / "market_cap.xlsx")
    if not path.exists():
        logger.error(f"market_cap.xlsx not found at path: {path}")
        raise FileNotFoundError(f"market_cap.xlsx not found at path: {path}")

    logger.info(f"Loading market cap data from {path}...")
    try:
        df = pd.read_excel(path)
    except Exception as e:
        logger.error(f"Failed to read Excel file at {path}: {e}")
        raise

    # Validation: Missing columns
    required_cols = [
        "company_id",
        "year",
        "market_cap_crore",
        "enterprise_value_crore",
        "pe_ratio",
        "pb_ratio",
        "ev_ebitda",
        "dividend_yield_pct",
    ]
    missing_cols = [col for col in required_cols if col not in df.columns]
    if missing_cols:
        err_msg = f"Missing required columns in market_cap.xlsx: {missing_cols}"
        logger.error(err_msg)
        raise ValueError(err_msg)

    # Clean company IDs
    df["company_id"] = df["company_id"].astype(str).str.strip().str.upper()

    # Convert year to integer
    df["year_int"] = df["year"].apply(extract_year_int)
    # Drop rows with invalid years for historical analysis
    df_clean = df.dropna(subset=["year_int"]).copy()
    df_clean["year_int"] = df_clean["year_int"].astype(int)

    # 5-year median PE (using up to the latest 5 years of data for each company)
    df_sorted = df_clean.sort_values(by=["company_id", "year_int"])
    df_5yr = df_sorted.groupby("company_id").tail(5)
    df_5yr_median = df_5yr.groupby("company_id")["pe_ratio"].median().reset_index()
    df_5yr_median.rename(columns={"pe_ratio": "5yr_median_PE"}, inplace=True)

    # Filter to latest year for each company
    latest_year_idx = df_clean.groupby("company_id")["year_int"].idxmax()
    df_latest = df_clean.loc[latest_year_idx].copy()

    # Merge latest with 5-year median PE
    df_merged = pd.merge(df_latest, df_5yr_median, on="company_id", how="left")

    # Validation: Duplicate company IDs
    if df_merged["company_id"].duplicated().any():
        dup_ids = (
            df_merged[df_merged["company_id"].duplicated()]["company_id"]
            .unique()
            .tolist()
        )
        logger.warning(
            f"Found duplicate company IDs in latest market cap records: {dup_ids}"
        )
        df_merged = df_merged.drop_duplicates(subset=["company_id"]).copy()

    # Validation: Missing market caps
    missing_mcap_count = df_merged["market_cap_crore"].isnull().sum()
    if missing_mcap_count > 0:
        logger.warning(
            f"Found {missing_mcap_count} companies with missing market cap in latest year."
        )

    logger.info(
        f"Loaded {len(df_merged)} company latest market cap records from Excel."
    )
    return df_merged


def load_company_master(db_path: Path | None = None) -> pd.DataFrame:
    """Loads id and company_name from companies table in SQLite DB."""
    path = db_path or DB_PATH
    logger.info(f"Connecting to database at {path} for company master...")
    conn = sqlite3.connect(str(path))
    try:
        df = pd.read_sql_query(
            "SELECT id AS company_id, company_name FROM companies", conn
        )
        df["company_id"] = df["company_id"].astype(str).str.strip().str.upper()
        logger.info(f"Loaded {len(df)} companies from database master.")
        return df
    except Exception as e:
        logger.error(f"Failed to query companies table: {e}")
        raise
    finally:
        conn.close()


def load_latest_ratios(db_path: Path | None = None) -> pd.DataFrame:
    """Loads latest-year financial ratios from SQLite DB."""
    path = db_path or DB_PATH
    conn = sqlite3.connect(str(path))
    try:
        # Load all ratios and parse year
        df = pd.read_sql_query("SELECT * FROM financial_ratios", conn)
        df["company_id"] = df["company_id"].astype(str).str.strip().str.upper()
        df["year_int"] = df["year"].apply(extract_year_int)
        df = df.dropna(subset=["year_int"]).copy()
        df["year_int"] = df["year_int"].astype(int)

        # Keep latest year only
        latest_idx = df.groupby("company_id")["year_int"].idxmax()
        df_latest = df.loc[latest_idx].copy()
        logger.info(
            f"Loaded {len(df_latest)} company latest financial ratio records from database."
        )
        return df_latest
    except Exception as e:
        logger.error(f"Failed to query financial_ratios: {e}")
        raise
    finally:
        conn.close()


def load_latest_financials(db_path: Path | None = None) -> pd.DataFrame:
    """Loads latest-year financial metrics (FCF) from SQLite DB."""
    # Since free_cash_flow_cr is inside financial_ratios, we can load it from load_latest_ratios
    df_ratios = load_latest_ratios(db_path)
    # We only need company_id and free_cash_flow_cr
    cols_to_keep = ["company_id", "free_cash_flow_cr"]
    # Check if column exists
    if "free_cash_flow_cr" not in df_ratios.columns:
        df_ratios["free_cash_flow_cr"] = None
    return df_ratios[cols_to_keep].copy()


def load_sector_data(db_path: Path | None = None) -> pd.DataFrame:
    """Loads company-sector mapping from SQLite DB."""
    path = db_path or DB_PATH
    conn = sqlite3.connect(str(path))
    try:
        df = pd.read_sql_query(
            "SELECT company_id, broad_sector AS sector FROM sectors", conn
        )
        df["company_id"] = df["company_id"].astype(str).str.strip().str.upper()
        logger.info(f"Loaded {len(df)} company sector records from database.")
        return df
    except Exception as e:
        logger.error(f"Failed to query sectors table: {e}")
        raise
    finally:
        conn.close()


def compute_valuation_metrics(
    df_mcap: pd.DataFrame,
    df_master: pd.DataFrame,
    df_sector: pd.DataFrame,
    df_fcf: pd.DataFrame,
) -> pd.DataFrame:
    """
    Merges datasets and calculates FCF Yield, Sector Median PE, Comparisons, and Flags.
    """
    logger.info("Merging datasets for valuation analysis...")
    # Begin with company master (92 companies) to ensure we cover the entire universe
    df_merged = pd.merge(df_master, df_sector, on="company_id", how="left")

    # Merge with market cap data
    df_merged = pd.merge(df_merged, df_mcap, on="company_id", how="left")

    # Merge with latest FCF
    df_merged = pd.merge(df_merged, df_fcf, on="company_id", how="left")

    logger.info("Computing FCF Yield %...")

    # Step 3: FCF Yield % = Free Cash Flow / Market Cap * 100
    def calc_fcf_yield(row: pd.Series) -> float:
        """Calc fcf yield."""
        fcf = row.get("free_cash_flow_cr")
        mcap = row.get("market_cap_crore")
        if pd.isnull(fcf) or pd.isnull(mcap) or mcap <= 0:
            return 0.0
        return round((float(fcf) / float(mcap)) * 100, 2)

    df_merged["FCF_yield_pct"] = df_merged.apply(calc_fcf_yield, axis=1)

    logger.info("Computing Sector Median PE...")
    # Step 4: Sector Median PE (Group by broad_sector, median of latest year PE)
    # Using pe_ratio from market_cap
    sector_medians = df_merged.groupby("sector")["pe_ratio"].median().reset_index()
    sector_medians.rename(columns={"pe_ratio": "sector_median_pe"}, inplace=True)

    df_merged = pd.merge(df_merged, sector_medians, on="sector", how="left")

    logger.info("Comparing Company PE against Sector Median...")

    # Step 5: PE vs Sector Median % = Company PE / Sector Median PE * 100
    def calc_pe_comparison(row: pd.Series) -> float | None:
        """Calc pe comparison."""
        pe = row.get("pe_ratio")
        sec_median = row.get("sector_median_pe")
        if pd.isnull(pe) or pd.isnull(sec_median) or sec_median <= 0:
            return None
        return round((float(pe) / float(sec_median)) * 100, 2)

    df_merged["PE_vs_sector_median_pct"] = df_merged.apply(calc_pe_comparison, axis=1)

    logger.info("Generating Valuation Flags...")

    # Step 6: Valuation Flags
    def get_valuation_flag(row: pd.Series) -> str:
        """Get valuation flag."""
        pe = row.get("pe_ratio")
        sec_median = row.get("sector_median_pe")
        if pd.isnull(pe) or pd.isnull(sec_median) or sec_median <= 0:
            return "Fair"  # Graceful fallback flag

        pe_vs_median = (float(pe) / float(sec_median)) * 100
        if pe_vs_median < 70.0:
            return "Discount"
        elif pe_vs_median > 150.0:
            return "Caution"
        else:
            return "Fair"

    df_merged["flag"] = df_merged.apply(get_valuation_flag, axis=1)

    # Rename and select final columns
    df_final = df_merged.copy()
    df_final.rename(
        columns={
            "pe_ratio": "PE",
            "pb_ratio": "PB",
            "ev_ebitda": "EV/EBITDA",
        },
        inplace=True,
    )

    # Fill NaN values in PE, PB, EV/EBITDA, 5yr_median_PE, and PE_vs_sector_median_pct with None (or leave for excel)
    # Ensure columns match final specification
    expected_cols = [
        "company_id",
        "company_name",
        "sector",
        "PE",
        "PB",
        "EV/EBITDA",
        "FCF_yield_pct",
        "5yr_median_PE",
        "PE_vs_sector_median_pct",
        "flag",
        "sector_median_pe",  # We keep this for flags.csv requirements
    ]
    # For columns that might not have merged or are completely empty
    for col in expected_cols:
        if col not in df_final.columns:
            df_final[col] = None

    # Step 7: Sort by Sector ascending, then Company Name
    df_final = (
        df_final[expected_cols]
        .sort_values(by=["sector", "company_name"])
        .reset_index(drop=True)
    )
    return df_final


def export_excel_report(df: pd.DataFrame, output_path: Path):
    """
    Exports the valuation summary to a styled Excel sheet under output/valuation_summary.xlsx.
    """
    logger.info(f"Exporting valuation summary to Excel at {output_path}...")
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # We don't need sector_median_pe in the Excel print to keep it clean, but let's see. Let's write the 10 final columns.
    excel_cols = [
        "company_id",
        "company_name",
        "sector",
        "PE",
        "PB",
        "EV/EBITDA",
        "FCF_yield_pct",
        "5yr_median_PE",
        "PE_vs_sector_median_pct",
        "flag",
    ]
    df_excel = df[excel_cols].copy()

    # Capitalize column headers for printing
    df_excel.columns = [
        "Company ID",
        "Company Name",
        "Sector",
        "P/E Ratio",
        "P/B Ratio",
        "EV/EBITDA",
        "FCF Yield %",
        "5Y Median P/E",
        "P/E vs. Sector Median %",
        "Valuation Flag",
    ]

    try:
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        writer = pd.ExcelWriter(output_path, engine="openpyxl")
        df_excel.to_excel(writer, sheet_name="Valuation Analysis", index=False)

        worksheet = writer.sheets["Valuation Analysis"]

        # Design system styles
        font_family = "Inter"
        header_fill = PatternFill(
            start_color="1E3A8A", end_color="1E3A8A", fill_type="solid"
        )  # Navy Blue
        header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
        data_font = Font(name=font_family, size=10)

        # Zebra striping and flag highlights
        zebra_fill = PatternFill(
            start_color="F8FAFC", end_color="F8FAFC", fill_type="solid"
        )  # light slate
        discount_fill = PatternFill(
            start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"
        )  # soft green
        caution_fill = PatternFill(
            start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"
        )  # soft red

        thin_border = Border(
            left=Side(style="thin", color="E2E8F0"),
            right=Side(style="thin", color="E2E8F0"),
            top=Side(style="thin", color="E2E8F0"),
            bottom=Side(style="thin", color="E2E8F0"),
        )

        # Style header
        for col_idx in range(1, len(df_excel.columns) + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            cell.border = thin_border

        worksheet.row_dimensions[1].height = 28

        # Style data rows
        for row_idx in range(2, len(df_excel) + 2):
            worksheet.row_dimensions[row_idx].height = 20
            is_zebra = row_idx % 2 == 0

            flag_val = str(worksheet.cell(row=row_idx, column=10).value).strip()

            for col_idx in range(1, len(df_excel.columns) + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.font = data_font
                cell.border = thin_border

                # Alignments
                if col_idx in [1, 3, 10]:  # ID, Sector, Flag
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif col_idx == 2:  # Name
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                else:  # Numeric ratios
                    cell.alignment = Alignment(horizontal="right", vertical="center")

                # Apply number formatting
                val = cell.value
                if val is not None and isinstance(val, (int, float)):
                    if col_idx in [4, 5, 6, 8]:  # Ratios
                        cell.number_format = "0.00"
                    elif col_idx in [7, 9]:  # Percentages
                        cell.number_format = (
                            "0.0%" if col_idx == 7 and val < 1.0 else "0.00"
                        )

                # Zebra striping
                if is_zebra:
                    cell.fill = zebra_fill

                # Highlight flags specifically
                if col_idx == 10:
                    if flag_val == "Discount":
                        cell.fill = discount_fill
                        cell.font = Font(
                            name=font_family, size=10, bold=True, color="166534"
                        )
                    elif flag_val == "Caution":
                        cell.fill = caution_fill
                        cell.font = Font(
                            name=font_family, size=10, bold=True, color="991B1B"
                        )

        # Freeze headers (Freeze Row 1, meaning we freeze at Row 2)
        worksheet.freeze_panes = "A2"

        # Auto column width
        for col in worksheet.columns:
            max_len = 0
            for cell in col:
                val = str(cell.value or "")
                max_len = max(max_len, len(val))
            col_letter = get_column_letter(col[0].column)
            worksheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

        writer.close()
        logger.info(f"Excel report successfully written to {output_path}")
    except Exception as e:
        logger.error(f"Failed to style and write Excel sheet: {e}")
        # Fallback to standard excel writing if openpyxl fails
        df_excel.to_excel(output_path, sheet_name="Valuation Analysis", index=False)
        logger.info(f"Fallback Excel report written to {output_path}")


def export_csv_flags_report(df: pd.DataFrame, output_path: Path):
    """
    Filters valuation summary for companies with flag != 'Fair' and
    writes output/valuation_flags.csv containing Company, Sector, PE, Sector Median, FCF Yield, Flag.
    """
    logger.info(f"Exporting valuation flags to CSV at {output_path}...")
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Filter flags
    df_flags = df[df["flag"].isin(["Discount", "Caution"])].copy()

    # Select and rename columns
    csv_cols = {
        "company_name": "Company",
        "sector": "Sector",
        "PE": "PE",
        "sector_median_pe": "Sector Median",
        "FCF_yield_pct": "FCF Yield",
        "flag": "Flag",
    }

    df_export = df_flags[list(csv_cols.keys())].rename(columns=csv_cols)

    # Save CSV
    df_export.to_csv(output_path, index=False)
    logger.info(
        f"CSV flags report written to {output_path} ({len(df_export)} flag records)"
    )


def run_valuation_pipeline(
    db_path: Path | None = None, raw_dir: Path | None = None
) -> pd.DataFrame:
    """Executes the complete valuation pipeline and exports reports."""
    start_time = time.time()
    logger.info("=" * 60)
    logger.info("        VALUATION ENGINE PIPELINE INITIALIZATION")
    logger.info("=" * 60)

    try:
        # 1. Load datasets
        df_mcap = load_market_cap(raw_dir)
        df_master = load_company_master(db_path)
        df_sector = load_sector_data(db_path)
        df_fcf = load_latest_financials(db_path)

        # 2. Process metrics
        df_final = compute_valuation_metrics(df_mcap, df_master, df_sector, df_fcf)

        # 3. Export reports
        summary_xlsx = OUTPUT_DIR / "valuation_summary.xlsx"
        flags_csv = OUTPUT_DIR / "valuation_flags.csv"

        export_excel_report(df_final, summary_xlsx)
        export_csv_flags_report(df_final, flags_csv)

        # 4. Validations
        row_count = len(df_final)
        duplicate_count = df_final["company_id"].duplicated().sum()
        missing_company_names = df_final["company_name"].isnull().sum()

        logger.info("-" * 60)
        logger.info("VALUATION VALIDATION SUMMARY:")
        logger.info(f"  Total Records: {row_count} (Expected: 92)")
        logger.info(f"  Duplicates:    {duplicate_count}")
        logger.info(f"  Missing Names: {missing_company_names}")

        # Log missing companies
        db_companies = set(df_master["company_id"].unique())
        processed_companies = set(df_final["company_id"].unique())
        missing_from_mc = db_companies - processed_companies
        if missing_from_mc:
            logger.warning(
                f"Companies in database missing from processed output: {missing_from_mc}"
            )

        runtime = round(time.time() - start_time, 2)
        logger.info(f"Valuation pipeline completed successfully in {runtime} seconds.")
        logger.info("=" * 60)

        return df_final

    except Exception as e:
        logger.critical(
            f"Valuation pipeline failed with critical error: {e}", exc_info=True
        )
        raise


if __name__ == "__main__":
    # Allow running standalone
    run_valuation_pipeline()
