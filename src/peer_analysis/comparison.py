"""
Peer Comparison Engine.
Coordinates data loading, grouping, percentile calculations, ranking, and Excel & PDF report generation.
"""

from __future__ import annotations

import datetime
import sqlite3
from pathlib import Path
from typing import Any, Dict, Optional, Tuple
# openpyxl for Excel report
import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
# ReportLab for PDF report
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import Image as RLImage
from reportlab.platypus import (Paragraph, SimpleDocTemplate, Spacer, Table,
                                TableStyle)

from src.config.settings import DB_PATH, OUTPUT_DIR
from src.peer_analysis.percentile import (calculate_sector_statistics,
                                          compute_percentiles)
from src.peer_analysis.summary import get_bottom_performers, get_top_performers
from src.utils.helpers import extract_year_int
from src.visualization.radar_chart import generate_peer_radar


def load_raw_ratios_data(db_path: Optional[Path] = None) -> pd.DataFrame:
    """
    Loads company details, sector info, and financial ratios from SQLite.
    """
    db_file = db_path or DB_PATH
    conn = sqlite3.connect(str(db_file))

    query = """
    SELECT 
        fr.company_id AS Company,
        s.broad_sector AS Sector,
        fr.return_on_equity_pct AS ROE,
        fr.return_on_capital_employed_pct AS ROCE,
        fr.revenue_cagr_5yr AS [Revenue CAGR],
        fr.pat_cagr_5yr AS [PAT CAGR],
        fr.debt_to_equity AS [Debt to Equity],
        fr.operating_profit_margin_pct AS [Operating Margin],
        fr.interest_coverage AS [Interest Coverage],
        fr.composite_quality_score AS [Composite Quality Score],
        fr.year
    FROM financial_ratios fr
    LEFT JOIN sectors s ON fr.company_id = s.company_id
    """
    try:
        df = pd.read_sql_query(query, conn)
    finally:
        conn.close()

    return df


def run_peer_analysis(
    db_path: Optional[Path] = None,
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    Executes the peer percentile and ranking analysis.

    Returns:
        Tuple of (peer_comparison_df, sector_stats_df, top_performers_df, bottom_performers_df)
    """
    df = load_raw_ratios_data(db_path)

    # Fill missing sectors with 'Unclassified'
    df["Sector"] = df["Sector"].fillna("Unclassified")

    # Keep only the latest year for each company
    df["year_int"] = df["year"].apply(extract_year_int)
    # Sort by year_int descending, then keep first company record
    df_latest = (
        df.sort_values(by="year_int", ascending=False)
        .drop_duplicates(subset=["Company"], keep="first")
        .copy()
    )

    # Percentile KPIs config: Column Name -> lower_is_better
    kpi_configs = {
        "ROE": False,
        "ROCE": False,
        "Revenue CAGR": False,
        "PAT CAGR": False,
        "Operating Margin": False,
        "Debt to Equity": True,  # Lower is better
        "Interest Coverage": False,
        "Composite Quality Score": False,
    }

    # Store calculated percentiles
    percentile_cols = {
        "ROE": "roe_percentile",
        "ROCE": "roce_percentile",
        "Revenue CAGR": "revenue_cagr_percentile",
        "PAT CAGR": "pat_cagr_percentile",
        "Operating Margin": "margin_percentile",
        "Debt to Equity": "de_percentile",
        "Interest Coverage": "interest_coverage_percentile",
        "Composite Quality Score": "quality_score_percentile",
    }

    # Initialize percentile columns
    for col in percentile_cols.values():
        df_latest[col] = 0.0

    # Group by Sector and compute percentiles & peer ranking
    ranked_groups = []

    for sector, group in df_latest.groupby("Sector"):
        # Calculate percentiles within the sector
        for kpi, pct_col in percentile_cols.items():
            group[pct_col] = compute_percentiles(
                group[kpi], lower_is_better=kpi_configs[kpi]
            )

        # Assign Peer Rank based on Composite Quality Score descending
        group_sorted = group.sort_values(
            by="Composite Quality Score", ascending=False
        ).copy()

        # Sequentially assign ranks (1, 2, 3...)
        group_sorted["Peer Rank"] = range(1, len(group_sorted) + 1)

        ranked_groups.append(group_sorted)

    df_ranked = pd.concat(ranked_groups).reset_index(drop=True)

    # 1. Create Peer Comparison Dataset
    peer_comparison = df_ranked.copy()

    # Map and rename columns to suggested names
    peer_comparison_export = pd.DataFrame(
        {
            "Company": peer_comparison["Company"],
            "Sector": peer_comparison["Sector"],
            "Peer Rank": peer_comparison["Peer Rank"],
            "Composite Score": peer_comparison["Composite Quality Score"],
            "ROE Percentile": peer_comparison["roe_percentile"],
            "ROCE Percentile": peer_comparison["roce_percentile"],
            "Revenue CAGR Percentile": peer_comparison["revenue_cagr_percentile"],
            "PAT CAGR Percentile": peer_comparison["pat_cagr_percentile"],
            "Debt to Equity Percentile": peer_comparison["de_percentile"],
            "Operating Margin Percentile": peer_comparison["margin_percentile"],
            "Interest Coverage Percentile": peer_comparison[
                "interest_coverage_percentile"
            ],
        }
    )

    # 2. Calculate Sector Statistics on raw KPIs
    kpis_to_stat = [
        "ROE",
        "ROCE",
        "Revenue CAGR",
        "PAT CAGR",
        "Operating Margin",
        "Debt to Equity",
        "Interest Coverage",
        "Composite Quality Score",
    ]
    sector_statistics = calculate_sector_statistics(df_latest, kpis_to_stat)

    # 3. Identify Top & Bottom Performers (Top 3 and Bottom 3 per Sector)
    top_performers = get_top_performers(df_ranked, n=3)
    bottom_performers = get_bottom_performers(df_ranked, n=3)

    # Clean summaries for export
    summary_cols = ["Company", "Sector", "Peer Rank", "Composite Quality Score"]
    top_performers_export = (
        top_performers[summary_cols]
        .rename(columns={"Composite Quality Score": "Composite Score"})
        .copy()
    )
    bottom_performers_export = (
        bottom_performers[summary_cols]
        .rename(columns={"Composite Quality Score": "Composite Score"})
        .copy()
    )

    # 4. Export Results
    csv_dir = OUTPUT_DIR / "csv"
    csv_dir.mkdir(parents=True, exist_ok=True)

    peer_comparison_export.to_csv(csv_dir / "peer_comparison.csv", index=False)
    sector_statistics.to_csv(csv_dir / "sector_statistics.csv", index=False)
    top_performers_export.to_csv(csv_dir / "top_performers.csv", index=False)
    bottom_performers_export.to_csv(csv_dir / "bottom_performers.csv", index=False)

    return (
        peer_comparison_export,
        sector_statistics,
        top_performers_export,
        bottom_performers_export,
    )


def get_raw_company_ratios(
    company_id: str, db_path: Optional[Path] = None
) -> Dict[str, Any]:
    """Retrieves raw KPIs for the latest year for the specified company."""
    df_all = load_raw_ratios_data(db_path)
    df_comp = df_all[df_all["Company"] == company_id].copy()
    if df_comp.empty:
        raise ValueError(f"Company {company_id} not found in database.")

    df_comp["year_int"] = df_comp["year"].apply(extract_year_int)
    latest_row = df_comp.sort_values(by="year_int", ascending=False).iloc[0]
    return latest_row.to_dict()


def determine_winner(
    kpi: str, val_a: Any, val_b: Any, company_a: str, company_b: str
) -> str:
    """Helper to determine winner for a specific KPI."""
    if pd.isnull(val_a) or val_a is None:
        return company_b
    if pd.isnull(val_b) or val_b is None:
        return company_a

    try:
        f_a = float(val_a)
        f_b = float(val_b)
    except (ValueError, TypeError):
        # Textual comparison (e.g. Sector, which doesn't have a winner)
        return "Tie"

    if kpi == "Debt to Equity":
        # Lower is better
        if f_a < f_b:
            return company_a
        elif f_b < f_a:
            return company_b
        else:
            return "Tie"
    else:
        # Higher is better
        if f_a > f_b:
            return company_a
        elif f_b > f_a:
            return company_b
        else:
            return "Tie"


def generate_executive_summary(
    company_a: str, company_b: str, a_data: Dict[str, Any], b_data: Dict[str, Any]
) -> str:
    """Generates an algorithmic, professional executive summary of the peer comparison."""
    # Compare growth metrics
    a_growth_wins = 0
    b_growth_wins = 0

    for kpi in ["Revenue CAGR", "PAT CAGR"]:
        w = determine_winner(
            kpi, a_data.get(kpi), b_data.get(kpi), company_a, company_b
        )
        if w == company_a:
            a_growth_wins += 1
        elif w == company_b:
            b_growth_wins += 1

    # Compare profitability
    a_profit_wins = 0
    b_profit_wins = 0
    for kpi in ["ROE", "ROCE", "Operating Margin"]:
        w = determine_winner(
            kpi, a_data.get(kpi), b_data.get(kpi), company_a, company_b
        )
        if w == company_a:
            a_profit_wins += 1
        elif w == company_b:
            b_profit_wins += 1

    # Composite Quality winner
    a_score = a_data.get("Composite Quality Score", 0.0)
    b_score = b_data.get("Composite Quality Score", 0.0)

    if a_score > b_score:
        overall_winner = company_a
        score_diff = a_score - b_score
        winner_text = f"{company_a} outperforms {company_b} overall with a stronger Composite Quality Score ({a_score:.1f} vs {b_score:.1f})."
    elif b_score > a_score:
        overall_winner = company_b
        score_diff = b_score - a_score
        winner_text = f"{company_b} outperforms {company_a} overall with a stronger Composite Quality Score ({b_score:.1f} vs {a_score:.1f})."
    else:
        overall_winner = "Tie"
        winner_text = f"Both {company_a} and {company_b} demonstrate identical Composite Quality Scores ({a_score:.1f})."

    summary_parts = [winner_text]

    # Describe growth strengths
    if a_growth_wins > b_growth_wins:
        summary_parts.append(
            f"{company_a} demonstrates superior performance in growth metrics (Revenue & PAT CAGRs)."
        )
    elif b_growth_wins > a_growth_wins:
        summary_parts.append(
            f"{company_b} demonstrates superior performance in growth metrics (Revenue & PAT CAGRs)."
        )
    else:
        summary_parts.append("Both companies demonstrate matching growth profiles.")

    # Describe profitability strengths
    if a_profit_wins > b_profit_wins:
        summary_parts.append(
            f"In terms of profitability, {company_a} leads in operational efficiency and returns on equity/capital."
        )
    elif b_profit_wins > a_profit_wins:
        summary_parts.append(
            f"In terms of profitability, {company_b} leads in operational efficiency and returns on equity/capital."
        )
    else:
        summary_parts.append(
            "They exhibit comparable operational profitability profiles."
        )

    return " ".join(summary_parts)


def generate_peer_comparison_report(
    company_a: str, company_b: str, db_path: Optional[Path] = None
) -> Tuple[Path, Path]:
    """
    Generates peer comparison reports:
    1. peer_comparison.xlsx (with openpyxl + embedded radar chart)
    2. peer_report.pdf (with reportlab + embedded radar chart)
    """
    db_file = db_path or DB_PATH

    # 1. Fetch raw ratios and verify
    a_data = get_raw_company_ratios(company_a, db_file)
    b_data = get_raw_company_ratios(company_b, db_file)

    # Get Peer Ranks
    peer_comp, _, _, _ = run_peer_analysis(db_file)
    try:
        a_rank = int(
            peer_comp[peer_comp["Company"] == company_a]["Peer Rank"].values[0]
        )
        a_sector = str(peer_comp[peer_comp["Company"] == company_a]["Sector"].values[0])
    except IndexError:
        a_rank = 1
        a_sector = "Unclassified"

    try:
        b_rank = int(
            peer_comp[peer_comp["Company"] == company_b]["Peer Rank"].values[0]
        )
        b_sector = str(peer_comp[peer_comp["Company"] == company_b]["Sector"].values[0])
    except IndexError:
        b_rank = 1
        b_sector = "Unclassified"

    # Ensure they are in the same sector
    sector_label = a_sector if a_sector == b_sector else f"{a_sector} vs {b_sector}"

    # 2. Build comparison data table
    kpis = [
        ("Sector", a_sector, b_sector, "@"),
        ("ROE (%)", a_data.get("ROE"), b_data.get("ROE"), "0.1f"),
        ("ROCE (%)", a_data.get("ROCE"), b_data.get("ROCE"), "0.1f"),
        (
            "Revenue CAGR (%)",
            a_data.get("Revenue CAGR"),
            b_data.get("Revenue CAGR"),
            "0.1f",
        ),
        ("PAT CAGR (%)", a_data.get("PAT CAGR"), b_data.get("PAT CAGR"), "0.1f"),
        (
            "Operating Margin (%)",
            a_data.get("Operating Margin"),
            b_data.get("Operating Margin"),
            "0.1f",
        ),
        (
            "Debt to Equity",
            a_data.get("Debt to Equity"),
            b_data.get("Debt to Equity"),
            "0.02f",
        ),
        (
            "Interest Coverage",
            a_data.get("Interest Coverage"),
            b_data.get("Interest Coverage"),
            "0.02f",
        ),
        (
            "Composite Score",
            a_data.get("Composite Quality Score"),
            b_data.get("Composite Quality Score"),
            "0.1f",
        ),
        ("Peer Rank", a_rank, b_rank, "d"),
    ]

    # 3. Generate Peer Radar Chart Image
    chart_dir = OUTPUT_DIR / "charts" / "peer"
    chart_dir.mkdir(parents=True, exist_ok=True)
    radar_png = chart_dir / f"{company_a}_vs_{company_b}_radar.png"
    generate_peer_radar(company_a, company_b, save_path=radar_png, db_path=db_file)

    # 4. Generate Executive Summary
    summary_text = generate_executive_summary(company_a, company_b, a_data, b_data)

    # ==========================================
    # A. EXCEL REPORT GENERATION (peer_comparison.xlsx)
    # ==========================================
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Peer Comparison"
    ws.views.sheetView[0].showGridLines = True

    # Styling variables
    navy_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    gold_fill = PatternFill(start_color="D4AF37", end_color="D4AF37", fill_type="solid")
    light_blue_fill = PatternFill(
        start_color="F2F6FA", end_color="F2F6FA", fill_type="solid"
    )
    white_fill = PatternFill(
        start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
    )
    zebra_fill = PatternFill(
        start_color="F9FBFD", end_color="F9FBFD", fill_type="solid"
    )

    title_font = Font(name="Calibri", size=16, bold=True, color="1B365D")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Calibri", size=11, bold=True, color="000000")
    regular_font = Font(name="Calibri", size=11, color="000000")
    italic_font = Font(name="Calibri", size=10, italic=True, color="595959")

    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    wrap_left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    # Title
    ws.cell(row=1, column=1, value="PEER COMPARISON REPORT").font = title_font
    ws.cell(
        row=2, column=1, value=f"Target: {company_a} vs {company_b} ({sector_label})"
    ).font = italic_font
    ws.cell(
        row=3,
        column=1,
        value=f"Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}",
    ).font = italic_font

    # Executive Summary Shaded Box
    ws.merge_cells("A5:D7")
    summary_cell = ws.cell(row=5, column=1, value=f"Executive Summary:\n{summary_text}")
    summary_cell.font = Font(name="Calibri", size=11, italic=True, color="000000")
    summary_cell.alignment = wrap_left_align
    summary_cell.fill = light_blue_fill

    # Apply border and fill to merged cells
    for r in range(5, 8):
        for c in range(1, 5):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            cell.fill = light_blue_fill

    # Headers
    headers = ["KPI Metric", f"{company_a}", f"{company_b}", "Winner"]
    ws.row_dimensions[9].height = 26
    for idx, h in enumerate(headers):
        c = idx + 1
        cell = ws.cell(row=9, column=c, value=h)
        cell.fill = navy_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    # Table Data
    for r_idx, (kpi, val_a, val_b, fmt) in enumerate(kpis):
        row = 10 + r_idx
        ws.row_dimensions[row].height = 20
        row_fill = zebra_fill if r_idx % 2 == 0 else white_fill

        # Winner calculation
        win_company = determine_winner(
            kpi.replace(" (%)", ""), val_a, val_b, company_a, company_b
        )

        # Cells
        c_kpi = ws.cell(row=row, column=1, value=kpi)
        c_val_a = ws.cell(row=row, column=2, value=val_a)
        c_val_b = ws.cell(row=row, column=3, value=val_b)
        c_win = ws.cell(row=row, column=4, value=win_company)

        for c in [c_kpi, c_val_a, c_val_b, c_win]:
            c.border = thin_border
            c.fill = row_fill
            c.font = regular_font

        c_kpi.alignment = left_align
        c_val_a.alignment = right_align
        c_val_b.alignment = right_align
        c_win.alignment = center_align

        # Apply formatting
        if fmt != "@" and fmt != "d":
            c_val_a.number_format = "0.0" if "1f" in fmt else "0.00"
            c_val_b.number_format = "0.0" if "1f" in fmt else "0.00"
        elif fmt == "d":
            c_val_a.number_format = "0"
            c_val_b.number_format = "0"
            c_val_a.alignment = center_align
            c_val_b.alignment = center_align

        # Highlight winner cell
        if win_company == company_a:
            c_win.font = Font(name="Calibri", size=11, bold=True, color="1B365D")
        elif win_company == company_b:
            c_win.font = Font(name="Calibri", size=11, bold=True, color="C00000")

    # Auto adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row >= 9:
                max_len = max(max_len, len(str(cell.value or "")))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 15)

    # Embed Radar Chart Image in Excel
    if radar_png.exists():
        img = openpyxl.drawing.image.Image(str(radar_png))
        img.width = 380
        img.height = 380
        ws.add_image(img, "F2")

    # Save Workbook
    excel_dir = OUTPUT_DIR / "reports"
    excel_dir.mkdir(parents=True, exist_ok=True)
    excel_path = excel_dir / "peer_comparison.xlsx"
    excel_path_root = OUTPUT_DIR / "peer_comparison.xlsx"
    wb.save(excel_path)
    wb.save(excel_path_root)

    # ==========================================
    # B. PDF REPORT GENERATION (peer_report.pdf)
    # ==========================================
    pdf_path = excel_dir / "peer_report.pdf"
    pdf_path_root = OUTPUT_DIR / "peer_report.pdf"

    # SimpleDocTemplate setup
    doc = SimpleDocTemplate(
        str(pdf_path),
        pagesize=letter,
        rightMargin=40,
        leftMargin=40,
        topMargin=40,
        bottomMargin=40,
    )

    styles = getSampleStyleSheet()

    # Custom Styles
    title_style = ParagraphStyle(
        "DocTitle",
        parent=styles["Heading1"],
        fontName="Helvetica-Bold",
        fontSize=20,
        leading=24,
        textColor=colors.HexColor("#1B365D"),
        spaceAfter=4,
    )

    subtitle_style = ParagraphStyle(
        "DocSubtitle",
        parent=styles["Normal"],
        fontName="Helvetica-Oblique",
        fontSize=10,
        leading=12,
        textColor=colors.HexColor("#595959"),
        spaceAfter=15,
    )

    summary_hdr_style = ParagraphStyle(
        "SummaryHeader",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=11,
        leading=14,
        textColor=colors.HexColor("#1B365D"),
        spaceAfter=5,
    )

    summary_body_style = ParagraphStyle(
        "SummaryBody",
        parent=styles["Normal"],
        fontName="Helvetica-Oblique",
        fontSize=10,
        leading=14,
        textColor=colors.HexColor("#333333"),
    )

    cell_style = ParagraphStyle(
        "TableCell",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=9,
        leading=11,
        textColor=colors.HexColor("#000000"),
    )

    cell_bold_style = ParagraphStyle(
        "TableCellBold", parent=cell_style, fontName="Helvetica-Bold"
    )

    cell_winner_a = ParagraphStyle(
        "TableCellWinnerA",
        parent=cell_style,
        fontName="Helvetica-Bold",
        textColor=colors.HexColor("#1B365D"),
    )

    cell_winner_b = ParagraphStyle(
        "TableCellWinnerB",
        parent=cell_style,
        fontName="Helvetica-Bold",
        textColor=colors.HexColor("#C00000"),
    )

    story = []

    # 1. Header Block
    story.append(Paragraph("NIFTY 100 FINANCIAL INTELLIGENCE PLATFORM", subtitle_style))
    story.append(
        Paragraph(f"PEER COMPARISON REPORT: {company_a} VS {company_b}", title_style)
    )
    story.append(
        Paragraph(
            f"Sector Universe: {sector_label}  |  Generated on: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}",
            subtitle_style,
        )
    )

    # 2. Shaded Executive Summary block
    summary_html = f"<b>Executive Summary:</b><br/>{summary_text}"
    summary_p = Paragraph(summary_html, summary_body_style)

    # Create single-cell table for shaded box background
    summary_box_table = Table([[summary_p]], colWidths=[530])
    summary_box_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F2F6FA")),
                ("BOX", (0, 0), (-1, -1), 1, colors.HexColor("#1B365D")),
                ("PADDING", (0, 0), (-1, -1), 10),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
            ]
        )
    )

    story.append(summary_box_table)
    story.append(Spacer(1, 15))

    # 3. Data Table Setup
    table_data = []

    # Table Header Row
    header_row = [
        Paragraph(
            "<b>KPI Metric</b>",
            ParagraphStyle("H1", parent=cell_bold_style, textColor=colors.white),
        ),
        Paragraph(
            f"<b>{company_a}</b>",
            ParagraphStyle("H2", parent=cell_bold_style, textColor=colors.white),
        ),
        Paragraph(
            f"<b>{company_b}</b>",
            ParagraphStyle("H3", parent=cell_bold_style, textColor=colors.white),
        ),
        Paragraph(
            "<b>Winner</b>",
            ParagraphStyle("H4", parent=cell_bold_style, textColor=colors.white),
        ),
    ]
    table_data.append(header_row)

    for r_idx, (kpi, val_a, val_b, fmt) in enumerate(kpis):
        win_company = determine_winner(
            kpi.replace(" (%)", ""), val_a, val_b, company_a, company_b
        )

        # Format string values
        str_a = (
            f"{val_a:.1f}%"
            if "1f" in fmt and kpi != "Sector"
            else (f"{val_a:.2f}" if "2f" in fmt else str(val_a))
        )
        str_b = (
            f"{val_b:.1f}%"
            if "1f" in fmt and kpi != "Sector"
            else (f"{val_b:.2f}" if "2f" in fmt else str(val_b))
        )

        if kpi == "Peer Rank":
            str_a = f"#{val_a}"
            str_b = f"#{val_b}"

        # Select style for winner cell
        if win_company == company_a:
            win_style = cell_winner_a
        elif win_company == company_b:
            win_style = cell_winner_b
        else:
            win_style = cell_style

        row = [
            Paragraph(kpi, cell_style),
            Paragraph(str_a, cell_style),
            Paragraph(str_b, cell_style),
            Paragraph(win_company, win_style),
        ]
        table_data.append(row)

    pdf_table = Table(table_data, colWidths=[180, 110, 110, 130])

    # Table Styling
    t_style = TableStyle(
        [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1B365D")),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("ALIGN", (1, 1), (2, -1), "RIGHT"),  # right align values
            ("ALIGN", (3, 0), (3, -1), "CENTER"),  # center winner column
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D9D9D9")),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]
    )

    # Apply zebra striping to rows
    for i in range(1, len(table_data)):
        bg = colors.HexColor("#F9FBFD") if i % 2 == 1 else colors.white
        t_style.add("BACKGROUND", (0, i), (-1, i), bg)

    pdf_table.setStyle(t_style)
    story.append(pdf_table)
    story.append(Spacer(1, 15))

    # 4. Insert Radar Chart Image (Center Aligned)
    if radar_png.exists():
        rl_img = RLImage(str(radar_png), width=310, height=310)
        rl_img.hAlign = "CENTER"
        story.append(rl_img)

    # Build Document
    doc.build(story)

    # Copy PDF to root output for accessibility
    import shutil

    shutil.copy(pdf_path, pdf_path_root)

    return excel_path, pdf_path


if __name__ == "__main__":
    # Test execution
    run_peer_analysis()
    try:
        generate_peer_comparison_report("INFY", "TCS")
        print("[+] Test reports generated successfully for INFY vs TCS!")
    except Exception as e:
        print(f"[-] Test report generation failed: {e}")
